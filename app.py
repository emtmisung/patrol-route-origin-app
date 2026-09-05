import base64
import calendar
import hmac
import html
import io
import math
import re
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, date, time as dtime, timedelta
from urllib.parse import quote

import folium
import pandas as pd
import qrcode
import requests
import streamlit as st
import streamlit.components.v1 as components
from streamlit_folium import st_folium

# ----------------------------------------------------------------------------
# 기본 설정
# ----------------------------------------------------------------------------
st.set_page_config(page_title="파세루 오리진 (FireSafe Route Origin)", page_icon="🚒", layout="wide")

GEOCODE_URL = "https://maps.apigw.ntruss.com/map-geocode/v2/geocode"
DIRECTIONS_URL = "https://maps.apigw.ntruss.com/map-direction/v1/driving"
NCP_KEY_ID = st.secrets.get("NCP_CLIENT_ID", "")
NCP_KEY = st.secrets.get("NCP_CLIENT_SECRET", "")
APP_PASSWORD = st.secrets.get("APP_PASSWORD", "")

AVG_SPEED_KMH = 35.0      # NCP 호출 실패 시에만 쓰는 비상 대체값(직선거리 보정)
ROAD_FACTOR = 1.3         # NCP 호출 실패 시에만 쓰는 비상 대체 보정계수

SAMPLE_XLSX = "seongju_patrol_coordinates_20.xlsx"


def ncp_headers():
    return {
        "x-ncp-apigw-api-key-id": NCP_KEY_ID,
        "x-ncp-apigw-api-key": NCP_KEY,
        "Accept": "application/json",
    }


def has_keys():
    return bool(NCP_KEY_ID) and bool(NCP_KEY)


# ----------------------------------------------------------------------------
# NCP API 호출
# ----------------------------------------------------------------------------
@st.cache_data(show_spinner=False, ttl=60 * 60 * 24)
def geocode_once(address: str):
    """주소 -> (lat, lng, 상태). 상태: "ok" | "not_found" | "error:..." """
    try:
        r = requests.get(
            GEOCODE_URL, params={"query": address}, headers=ncp_headers(), timeout=10
        )
        if r.status_code != 200:
            return None, None, f"error:HTTP {r.status_code}"
        data = r.json()
        addrs = data.get("addresses") or []
        if addrs:
            a = addrs[0]
            return float(a["y"]), float(a["x"]), "ok"
        return None, None, "not_found"
    except Exception as e:
        return None, None, f"error:{type(e).__name__}"


def geocode_address(address: str):
    """기존 호출부 호환용 — (lat, lng)만 반환."""
    lat, lng, _ = geocode_once(address)
    return lat, lng


def address_variants(address: str, name: str = ""):
    """지오코딩이 실패했을 때 순서대로 다시 시도할 주소 후보들을 만든다.

    행정리('성산1리')는 지오코딩이 인식하지 못하는 경우가 많아
    법정리('성산리')로 바꾸는 것이 가장 중요한 보정이다.
    """
    address = (address or "").strip()
    cands = []

    def add(v, why):
        v = re.sub(r"\s+", " ", (v or "")).strip()
        if v and all(v != c[0] for c in cands):
            cands.append((v, why))

    add(address, "원본 주소")

    # 1) 행정리 번호 제거: 성산1리 → 성산리, 경산8리 → 경산리
    v1 = re.sub(r"([가-힣]+?)\d+리(?=\s|$)", r"\1리", address)
    add(v1, "행정리→법정리 (성산1리→성산리)")

    # 2) 괄호와 그 안의 내용 제거
    v2 = re.sub(r"\([^)]*\)", " ", v1)
    add(v2, "괄호 제거")

    # 3) 지번의 부번 제거: 540-1 → 540
    v3 = re.sub(r"(\d+)-\d+(?=\s|$)", r"\1", v2)
    add(v3, "지번 부번 제거 (540-1→540)")

    # 4) 번지 자체를 떼고 리(동) 중심으로: … 성산리 1805 → … 성산리
    v4 = re.sub(r"\s+\d+(-\d+)?\s*$", "", v3)
    add(v4, "번지 제외 (리·동 중심 좌표)")

    # 5) 대상명 안 괄호에 들어 있는 주소를 활용: 차동골 마을회관 (성주읍 성산1리 1805)
    m = re.search(r"\(([^)]*)\)", name or "")
    if m:
        inner = re.sub(r"([가-힣]+?)\d+리(?=\s|$)", r"\1리", m.group(1))
        add(f"경상북도 성주군 {inner}", "대상명 속 주소 사용")
        add(re.sub(r"\s+\d+(-\d+)?\s*$", "", f"경상북도 성주군 {inner}"), "대상명 속 주소(번지 제외)")

    return cands


def geocode_with_fallback(address: str, name: str = "", on_call=None):
    """여러 주소 형태로 순차 시도. 반환: (lat, lng, 성공에 쓴 주소, 방법, 시도내역)"""
    tried = []
    for query, why in address_variants(address, name):
        lat, lng, status = geocode_once(query)
        if on_call:
            on_call()
        tried.append(f"{why}: {query} → {status}")
        if status == "ok":
            return lat, lng, query, why, tried
    return None, None, None, None, tried


@st.cache_data(show_spinner=False, ttl=60 * 60 * 24)
def road_route(o_lat, o_lng, d_lat, d_lng):
    """실도로 거리(km)/시간(분)/경로좌표 반환. 실패 시 None 튜플."""
    try:
        r = requests.get(
            DIRECTIONS_URL,
            params={
                "start": f"{o_lng},{o_lat}",
                "goal": f"{d_lng},{d_lat}",
                "option": "trafast",
            },
            headers=ncp_headers(),
            timeout=10,
        )
        data = r.json()
        route = data.get("route", {})
        for key in ("trafast", "traoptimal", "tracomfort"):
            if key in route and route[key]:
                summ = route[key][0]["summary"]
                path = route[key][0].get("path", [])
                return (
                    summ["distance"] / 1000.0,
                    summ["duration"] / 60000.0,
                    [(p[1], p[0]) for p in path],
                )
        return None, None, None
    except Exception:
        return None, None, None


def haversine_km(a, b):
    lat1, lng1 = a
    lat2, lng2 = b
    R = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lng2 - lng1)
    x = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlmb / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(x), math.sqrt(1 - x))


# ----------------------------------------------------------------------------
# 한글(hwpx) 표 파싱 — 데모(웹 프로토타입)와 동일한 방식
# ----------------------------------------------------------------------------
HEADER_WORDS = re.compile(r"^(연번|no\.?|번호|구분|이름|명칭|대상명|대상명주소|주소|정제_주소|비고)$", re.I)


def _clean_xml_text(s: str) -> str:
    s = re.sub(r"<[^>]+>", "", s)
    return (s.replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")
             .replace("&quot;", '"').replace("&apos;", "'").strip())


def parse_hwpx(file_bytes: bytes):
    """hwpx 안의 표(또는 문단)를 읽어 DataFrame으로 반환."""
    rows = []
    with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
        for name in zf.namelist():
            if not re.search(r"Contents/section\d*\.xml$", name, re.I):
                continue
            xml = zf.read(name).decode("utf-8", errors="ignore")
            table_rows = re.findall(r"<hp:tr[\s>][\s\S]*?</hp:tr>", xml)
            if table_rows:
                for row_xml in table_rows:
                    cells = re.findall(r"<hp:tc[\s>][\s\S]*?</hp:tc>", row_xml)
                    cols = [_clean_xml_text("".join(re.findall(r"<hp:t[^>]*>([\s\S]*?)</hp:t>", c)))
                            for c in cells]
                    if any(cols):
                        rows.append(cols)
            else:
                for chunk in xml.split("<hp:p")[1:]:
                    text = _clean_xml_text("".join(re.findall(r"<hp:t[^>]*>([\s\S]*?)</hp:t>", chunk)))
                    if text:
                        rows.append([text])

    rows = [r for r in rows if r and not HEADER_WORDS.match((r[0] or "").strip())]
    if not rows:
        return None

    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    default_names = ["연번", "주소지", "비고", "정제_주소", "위도(Latitude)", "경도(Longitude)"]
    cols = default_names[:width] + [f"열{i}" for i in range(len(default_names) + 1, width + 1)]
    return pd.DataFrame(rows, columns=cols[:width])


# ----------------------------------------------------------------------------
# 경로 편성 알고리즘 (최근접 이웃 기반, 소방서 출발/복귀)
# 거리·시간 판단은 전부 NCP Directions5의 실제 도로거리를 사용한다.
# (직선거리는 API 호출이 실패했을 때만 비상 대체값으로 쓰인다)
# ----------------------------------------------------------------------------
def real_leg(a, b, on_call=None):
    """a, b: dict(lat, lng). 실도로 거리(km)/시간(분) 반환 (실패 시 직선거리 보정값)."""
    km, mins, _ = road_route(a["lat"], a["lng"], b["lat"], b["lng"])
    if on_call:
        on_call()
    if km is None:
        km = haversine_km((a["lat"], a["lng"]), (b["lat"], b["lng"])) * ROAD_FACTOR
        mins = km / AVG_SPEED_KMH * 60
    return km, mins


def nearest_by_straight_line(cur, candidates, k):
    """직선거리로 가까운 순 k개만 추린다. (실제 API 호출 횟수를 줄이기 위한 1차 필터)

    도로망은 직선거리와 순서가 크게 다르지 않으므로, 가까운 후보 몇 개만
    실제 도로거리로 확인해도 결과는 거의 동일하면서 API 호출은 크게 줄어든다.
    """
    if k <= 0 or k >= len(candidates):
        return candidates
    ranked = sorted(
        candidates,
        key=lambda p: haversine_km((cur["lat"], cur["lng"]), (p["lat"], p["lng"])),
    )
    return ranked[:k]


def build_routes(points, station, mode, max_per_route, seg_max_km, seg_max_min,
                 target_min_high, max_routes_cap, basis="distance", on_call=None,
                 candidate_k=5, should_stop=None, service_min_per_stop=0):
    """points: list of dict(name, address, lat, lng)
    반환: routes(list of list of point dict), unassigned(장거리/미배정)

    mode:
      "segment"     — 구간당 거리·시간 제한
      "target_time" — 노선 전체 왕복 목표시간 제한
      "fixed"       — 노선당 구간 수(max_per_route)를 그대로 채움 (노선 수 = 상한까지)
    basis: "distance"(거리 기준) | "time"(소요시간 기준)
    candidate_k: 다음 지점 후보를 직선거리로 몇 개까지 좁혀서 실제 API로 확인할지 (0=전수)
    should_stop: 호출 한도 초과 등으로 중단해야 하는지 판단하는 함수.
                 중단되면 그때까지 편성된 노선만 반환한다(진행분 보존).
    """
    remaining = points[:]
    routes = []

    guard = 0
    while remaining and guard < 500:
        if should_stop and should_stop():
            break
        guard += 1
        cur = station
        route = []
        acc_min = 0.0

        while remaining:
            if should_stop and should_stop():
                break
            # 1차: 직선거리로 후보 좁히기 → 2차: 좁혀진 후보만 실도로 거리/시간 확인
            candidates = nearest_by_straight_line(cur, remaining, candidate_k)
            legs = [(p, *real_leg(cur, p, on_call)) for p in candidates]
            legs.sort(key=(lambda t: t[2]) if basis == "time" else (lambda t: t[1]))
            nxt, leg_km, leg_min = legs[0]

            # 노선의 첫 지점은 제한값을 적용하지 않는다.
            # (소방서에서 가장 가까운 대상까지의 거리가 이미 제한값보다 크면
            #  어떤 노선도 못 만들고 전부 '장거리'로 빠지는 문제를 막기 위함)
            first_stop = not route

            if mode == "fixed":
                if len(route) >= max_per_route:
                    break
            elif mode == "segment":
                if not first_stop and (leg_km > seg_max_km or leg_min > seg_max_min):
                    break
                if len(route) >= max_per_route:
                    break
            else:  # target_time
                back_km, back_min = real_leg(nxt, station, on_call)
                projected = acc_min + leg_min + service_min_per_stop + back_min
                if not first_stop and projected > target_min_high:
                    break
                if len(route) >= max_per_route:
                    break

            route.append(nxt)
            acc_min += leg_min + service_min_per_stop
            cur = nxt
            remaining.remove(nxt)

        if not route:
            # 어떤 조건도 만족 못하는 경우(예: 첫 지점부터 원거리) -> 강제 배정 방지, 장거리로 이관
            break
        routes.append(route)

        if max_routes_cap and len(routes) >= max_routes_cap and remaining:
            # 노선 수 상한 도달 -> 남은 지점은 마지막 노선에 최대한 이어붙임(완화)
            for p in remaining[:]:
                route.append(p)
                remaining.remove(p)
            break

    return routes, remaining


def allocate_hydrants_to_members(points, station, members):
    """개인별 개수 차이를 1개 이하로 유지하면서 인접 구역으로 배정한다.

    같은 차량의 팀원을 연속 배치한 뒤 센터 기준 방위각으로 정렬한 소화전을
    연속 구간으로 나눠, 같은 차량 팀원들의 담당 구역도 서로 가깝게 만든다.
    """
    if not points or not members:
        return points

    ordered_members = sorted(members, key=lambda m: (m["vehicle_no"], m["order"]))
    ordered_points = sorted(
        points,
        key=lambda p: (
            math.atan2(p["lat"] - station["lat"], p["lng"] - station["lng"]),
            haversine_km((station["lat"], station["lng"]), (p["lat"], p["lng"])),
        ),
    )
    base, extra = divmod(len(ordered_points), len(ordered_members))
    assigned = []
    cursor = 0
    for index, member in enumerate(ordered_members):
        count = base + (1 if index < extra else 0)
        for point in ordered_points[cursor:cursor + count]:
            assigned.append({
                **point,
                "assigned_to": member["name"],
                "vehicle_no": member["vehicle_no"],
            })
        cursor += count
    return assigned


def separate_long_distance(points, station, threshold_km, on_call=None, save_calls=True,
                           should_stop=None):
    """소방서에서 실도로거리가 기준을 넘는 대상을 분리한다.

    save_calls=True면 직선거리 추정값이 기준에서 충분히 멀리 떨어진(애매하지 않은)
    대상은 API를 호출하지 않고 추정값으로 판정해 호출 횟수를 줄인다.
    """
    normal, far = [], []
    for p in points:
        straight = haversine_km((station["lat"], station["lng"]), (p["lat"], p["lng"]))
        est = straight * ROAD_FACTOR

        if should_stop and should_stop():
            # 한도 초과 — 남은 대상은 추정값으로 분류하고 API 호출은 더 하지 않는다
            (far if est > threshold_km else normal).append(
                {**p, "도로거리_km": round(est, 1)} if est > threshold_km else p
            )
            continue

        if save_calls and est < threshold_km * 0.7:
            normal.append(p)          # 확실히 가까움 — API 호출 생략
            continue
        if save_calls and est > threshold_km * 1.5:
            far.append({**p, "도로거리_km": round(est, 1)})  # 확실히 멂 — 추정값 사용
            continue

        km, _ = real_leg(station, p, on_call)   # 애매한 구간만 실제 도로거리로 확인
        if km > threshold_km:
            far.append({**p, "도로거리_km": round(km, 1)})
        else:
            normal.append(p)
    return normal, far


def separate_long_time(points, station, threshold_min, delegate_to, on_call=None,
                       should_stop=None):
    """센터 기준 실제 편도시간으로 계절순찰 대상과 원거리 위임 대상을 나눈다."""
    normal, far = [], []
    for point in points:
        if should_stop and should_stop():
            straight_km = haversine_km(
                (station["lat"], station["lng"]), (point["lat"], point["lng"])
            )
            est_km = straight_km * ROAD_FACTOR
            est_min = est_km / AVG_SPEED_KMH * 60
            far.append({
                **point,
                "도로거리_km": round(est_km, 1),
                "편도시간_분": round(est_min),
                "권장수행": f"{delegate_to} (API 한도 도달로 재확인 필요)",
            })
            continue
        km, mins = real_leg(station, point, on_call)
        if mins > threshold_min:
            far.append({
                **point,
                "도로거리_km": round(km, 1),
                "편도시간_분": round(mins),
                "권장수행": delegate_to,
            })
        else:
            normal.append(point)
    return normal, far


# ----------------------------------------------------------------------------
# UI — 파세루 데모(웹 프로토타입)와 같은 카드+칩 스타일
# ----------------------------------------------------------------------------
PASERU_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700;800&family=Noto+Serif+KR:wght@600;700&family=IBM+Plex+Mono:wght@500;700&display=swap');

:root{
  --accent:#a33a3f; --accent-button:#b8464b; --accent-hover:#8f3035; --accent-soft:#f7e9ea;
  --navy:#17263a; --ink:#263442; --muted:#344054;
  --line:#cbd3dd; --divider:#d8dee7; --surface:#ffffff; --bg:#f7f8fa;
  --focus:#d8898d;
  color-scheme: light;   /* 휴대폰 다크모드에서도 밝은 화면으로 고정 */
}
.stApp{ background:var(--bg); color:var(--ink); font-size:14px; line-height:1.55; }

/* 휴대폰 다크모드에서 '흰 배경 + 흰 글씨'가 되는 문제를 막기 위해
   본문 글자색을 어두운 색으로 명시적으로 고정한다. */
.stApp, .stApp p, .stApp span, .stApp label, .stApp li, .stApp div,
.stMarkdown, .stMarkdown *, [data-testid="stWidgetLabel"] *,
[data-testid="stMetricLabel"] *, [data-testid="stMetricValue"],
[data-testid="stCaptionContainer"], [data-testid="stCaptionContainer"] *{
  color:var(--ink);
}
[data-testid="stCaptionContainer"], [data-testid="stCaptionContainer"] *,
[data-testid="stMetricLabel"] *{ color: var(--muted) !important; }

/* 알림 박스(노란색·파란색 등) 안 글씨도 항상 검정 계열로 */
div[data-testid="stAlert"], div[data-testid="stAlert"] *,
div[data-testid="stNotification"], div[data-testid="stNotification"] *{
  color:var(--ink) !important;
}

/* 입력창·표를 밝은 배경 + 어두운 글씨로 고정 */
input, textarea, select,
[data-baseweb="input"] input, [data-baseweb="base-input"] input,
[data-baseweb="select"] div{
  background-color:#ffffff !important; color:var(--ink) !important;
  border-color:#aab4c1 !important;
}
[data-testid="stDataFrame"], [data-testid="stDataEditor"],
[data-testid="stTable"]{ background:#ffffff !important; }

html, body, [class*="css"], .stMarkdown, .stTextInput, .stNumberInput{
  font-family:'Noto Sans KR', -apple-system, 'Malgun Gothic', sans-serif;
}
h1, h2, h3, h4, h5, h6{
  font-family:'Noto Serif KR', serif !important; color:var(--navy) !important; font-weight:700 !important;
}
h1{ font-size:32px !important; }
h2{ font-size:24px !important; }
h3{ font-size:20px !important; }
.block-container{ padding-top: 2.2rem; max-width: 1180px; }

/* ---- 카드 컨테이너(border=True) ---- */
div[data-testid="stVerticalBlockBorderWrapper"]{
  background: var(--surface);
  border-radius: 14px !important;
  border:1px solid var(--line) !important;
  border-left:4px solid var(--accent) !important;
  box-shadow:0 2px 10px rgba(23,38,58,.06);
  padding:12px 18px 16px;
  margin-bottom:8px;
}

/* ---- 카드 제목 + 번호 뱃지 ---- */
.paseru-card-title{
  display:flex; align-items:center; gap:9px;
  font-family:'Noto Serif KR', serif; font-size:18px; font-weight:700; color:var(--navy);
  margin: 2px 0 10px;
}
.paseru-step{
  display:inline-flex; align-items:center; justify-content:center;
  width:23px; height:23px; border-radius:50%;
  background:var(--accent); color:#fff !important;
  font-family:'IBM Plex Mono', monospace; font-size:12px; font-weight:700; flex:none;
}
.paseru-sub{ font-weight:650; font-size:14px; color:#22324a; margin:14px 0 6px; }
.paseru-eyebrow{
  font-family:'IBM Plex Mono', monospace; font-size:12px; letter-spacing:.08em;
  text-transform:uppercase; color:var(--accent); font-weight:700; margin-bottom:2px;
}

/* ---- 선택 칩: 미선택도 선명한 테두리, 선택 시 차분한 딥 레드 ---- */
button[data-variant="pills"]{
  border-radius: 999px !important;
  border: 1px solid var(--line) !important;
  background:#f4f6f8 !important;
  color: var(--ink) !important;
  border-color:#aab4c1 !important;
  font-weight:600 !important;
  padding: 0.42em 1.05em !important;
  transition: background .12s, border-color .12s, color .12s;
}
button[data-variant="pills"]:hover{ background:#f7e9ea !important; border-color:var(--accent) !important; }
button[data-variant="pills"][data-selected="true"],
button[data-variant="pills"][aria-checked="true"],
button[data-variant="pills"][aria-pressed="true"]{
  background: var(--accent) !important;
  border-color: var(--accent) !important;
  color: #ffffff !important;
  font-weight: 700 !important;
  box-shadow:0 4px 12px -7px rgba(163,58,63,.42);
}
button[data-variant="pills"][data-selected="true"] *,
button[data-variant="pills"][aria-checked="true"] *,
button[data-variant="pills"][aria-pressed="true"] *{ color:#ffffff !important; }
div[data-testid="stButtonGroup"]{ gap: 8px !important; }

/* ---- 버튼 ---- */
div.stButton > button, .stDownloadButton > button, div.stFormSubmitter > button{
  background-color:var(--accent-button) !important;
  color:#fff !important; border:none !important; border-radius:10px !important;
  font-weight:700 !important; padding:0.98em 1.1em !important;
  box-shadow:0 5px 14px -8px rgba(143,48,53,.48);
}
div.stButton > button:hover, .stDownloadButton > button:hover{ background-color: var(--accent-hover) !important; }
/* 전역 글자색 규칙보다 우선해 주요 빨간 버튼의 글자를 항상 흰색으로 표시 */
div.stButton > button:not([kind="secondary"]),
div.stButton > button:not([kind="secondary"]) *,
.stDownloadButton > button, .stDownloadButton > button *,
div.stFormSubmitter > button, div.stFormSubmitter > button *{
  color:#ffffff !important;
  -webkit-text-fill-color:#ffffff !important;
  opacity:1 !important;
}
div.stButton > button[kind="secondary"]{
  background:#ffffff !important; color:var(--ink) !important;
  border:1px solid #aab4c1 !important; box-shadow:none !important;
}
button:focus-visible, input:focus-visible, textarea:focus-visible,
[tabindex]:focus-visible{ outline:2px solid var(--focus) !important; outline-offset:2px; }

/* ---- 결과 영역 ---- */
div[data-testid="stExpander"]{
  border:1px solid var(--line) !important; border-radius:12px !important;
  background:var(--surface); overflow:hidden;
  box-shadow:0 1px 6px rgba(23,38,58,.04);
}

/* ---- 대비 강화: 지표·표·라벨이 흐리게 보이지 않도록 ---- */
div[data-testid="stMetricValue"]{
  font-family:'IBM Plex Mono', monospace;
  color:#315d78 !important; font-weight:700 !important;
}
div[data-testid="stMetricLabel"], div[data-testid="stMetricLabel"] *{
  color:var(--ink) !important; font-weight:600 !important;
}
[data-testid="stWidgetLabel"] p, [data-testid="stWidgetLabel"] label{
  color:var(--ink) !important; font-weight:600 !important;
}
[data-testid="stCaptionContainer"], [data-testid="stCaptionContainer"] *{
  color:var(--muted) !important;
  -webkit-text-fill-color:var(--muted) !important;
  opacity:1 !important;
  font-weight:500 !important;
}
/* 안내문·업로더 보조문이 연한 회색으로 흐려지지 않도록 대비 확보 */
.stApp small, .stApp small *,
[data-testid="stFileUploader"] small,
[data-testid="stFileUploader"] small *,
[data-testid="stFileUploaderDropzoneInstructions"],
[data-testid="stFileUploaderDropzoneInstructions"] *{
  color:var(--muted) !important;
  -webkit-text-fill-color:var(--muted) !important;
  opacity:1 !important;
}
/* 표(데이터프레임·편집표) 글씨와 테두리를 진하게 */
[data-testid="stDataFrame"] *, [data-testid="stDataEditor"] *{
  color:var(--ink) !important;
}
[data-testid="stDataFrame"], [data-testid="stDataEditor"]{
  border:1px solid var(--line) !important; border-radius:8px;
}
[data-testid="stDataFrame"] [role="columnheader"],
[data-testid="stDataEditor"] [role="columnheader"]{
  background:#eef2f6 !important; color:var(--navy) !important; font-weight:700 !important;
}
/* 알림 박스에 색 띠를 넣어 눈에 잘 띄게 */
div[data-testid="stAlert"]{
  border:1px solid #a9bfce !important;
  border-left:4px solid #557f99 !important; border-radius:10px !important;
  background:#eef4f8 !important;
}
.paseru-sub{ color:#22324a !important; }

/* 완료된 핵심 작업은 기존 실행 버튼 자리에 초록색 상태 버튼처럼 표시 */
.paseru-complete-action{
  width:100%; padding:.72rem 1rem; border-radius:10px;
  background:#2f6b49; border:1px solid #24563b;
  color:#ffffff !important; text-align:center; font-weight:750;
  box-shadow:0 4px 12px -8px rgba(36,86,59,.55);
}
.paseru-complete-action *{ color:#ffffff !important; }

/* ---- 내비게이션 버튼: 링크 기본색(파랑)에 밀리지 않도록 클래스로 고정 ---- */
a.paseru-navbtn, a.paseru-navbtn:link, a.paseru-navbtn:visited,
a.paseru-navbtn:hover, a.paseru-navbtn:active,
a.paseru-navbtn *{
  color:#ffffff !important;
  text-decoration:none !important;
}
a.paseru-navbtn{
  display:inline-block; padding:12px 16px; border-radius:10px;
  font-weight:700 !important; font-size:14px; margin:4px 8px 4px 0;
  box-shadow:0 3px 10px -4px rgba(0,0,0,.35);
}
a.paseru-navbtn.nav-and{ background:#03C75A !important; }   /* 안드로이드 (네이버 초록) */
a.paseru-navbtn.nav-ios{ background:#0a8f45 !important; }   /* 아이폰 */
a.paseru-navbtn.nav-pc{  background:#2563eb !important; }   /* PC 웹 (밝은 파랑) */
</style>
"""
st.markdown(PASERU_CSS, unsafe_allow_html=True)


def stop_label(name, address):
    """대상명 + 주소 표기. 이름 안에 이미 주소(또는 번지)가 들어 있으면 중복 표기하지 않는다."""
    name = (name or "").strip()
    address = (address or "").strip()
    if not address or address == name:
        return name
    # "경상북도 성주군 월항면 인촌1리 606-1" -> 뒤쪽 핵심부("인촌1리 606-1")가 이름에 있으면 생략
    tail = " ".join(address.split()[-2:])
    if tail and tail in name:
        return name
    if address in name:
        return name
    return f"{name} ({address})"


def kakao_url(name, lat, lng):
    """카카오맵 길안내 링크 (공백·괄호가 있어도 깨지지 않도록 인코딩)."""
    return ("https://map.kakao.com/link/to/"
            f"{quote(str(name), safe='')},{lat},{lng}")


KAKAO_MAX_VIA = 5  # 카카오맵 자동차 길찾기 URL이 지원하는 경유지 최대 개수


def kakao_route_url(origin, destinations):
    """카카오맵 자동차 길찾기 링크를 만든다.

    origin은 출발지, destinations의 마지막 항목은 목적지이며 그 앞 항목은
    경유지로 전달된다. destinations는 최대 6개(경유지 5 + 목적지)다.
    """
    if not destinations:
        return ""

    def place(p):
        name = quote(str(p["name"]), safe="")
        return f"{name},{float(p['lat']):.7f},{float(p['lng']):.7f}"

    points = [origin] + list(destinations)
    return "https://map.kakao.com/link/by/car/" + "/".join(place(p) for p in points)


def kakao_route_links(station, legs):
    """소방서 → 경유지 순서 → 소방서로 돌아오는 카카오맵 링크 목록.

    경유지가 5개를 넘는 긴 노선은 앞 구간의 마지막 목적지를 다음 구간의
    출발지로 이어서 분할한다.
    반환: [(URL, 출발지, 구간 목적지 목록), ...]
    """
    stops = [{"name": lg["to"], "lat": lg["lat"], "lng": lg["lng"]} for lg in legs]
    if not stops:
        return []

    remaining = stops + [station]
    origin = station
    links = []
    max_destinations = KAKAO_MAX_VIA + 1
    while remaining:
        destinations = remaining[:max_destinations]
        links.append((kakao_route_url(origin, destinations), origin, destinations))
        remaining = remaining[max_destinations:]
        if remaining:
            origin = destinations[-1]
    return links


def make_qr_png(data):
    """링크를 휴대폰으로 넘길 수 있는 QR코드 PNG 바이트로 만든다."""
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=8,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    image = qr.make_image(fill_color="black", back_color="white")
    buffer = io.BytesIO()
    image.save(buffer, format="PNG")
    return buffer.getvalue()


def build_qr_zip(station, route_results):
    """모든 노선의 카카오맵 QR PNG와 경로 목록 엑셀을 ZIP으로 묶는다."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for rr in route_results:
            links = kakao_route_links(station, rr["legs"])
            for li, (url, origin, destinations) in enumerate(links, start=1):
                suffix = "" if len(links) == 1 else f"_구간{li}"
                zf.writestr(f"노선_{rr['route_no']}{suffix}_QR.png", make_qr_png(url))
        zf.writestr("노선별_경로와_링크.xlsx", build_route_links_excel(station, route_results))
    return buffer.getvalue()


def build_route_links_excel(station, route_results):
    """노선 순서와 클릭 가능한 카카오맵 링크를 엑셀로 만든다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "노선별 경로와 링크"
    ws.sheet_view.showGridLines = False

    headers = ["노선", "구간", "출발지", "경유지 및 목적지 순서", "거리(km)", "시간(분)", "카카오맵"]
    ws.append(headers)

    for rr in route_results:
        links = kakao_route_links(station, rr["legs"])
        for li, (url, origin, destinations) in enumerate(links, start=1):
            sequence = " → ".join([origin["name"]] + [p["name"] for p in destinations])
            ws.append([
                rr["route_no"],
                li if len(links) > 1 else 1,
                origin["name"],
                sequence,
                round(rr["total_km"], 1),
                round(rr["total_min"]),
                "카카오맵에서 열기",
            ])
            link_cell = ws.cell(row=ws.max_row, column=7)
            link_cell.hyperlink = url
            link_cell.style = "Hyperlink"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    bottom_border = Border(bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = bottom_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        row[0].alignment = Alignment(horizontal="center", vertical="center")
        row[1].alignment = Alignment(horizontal="center", vertical="center")
        row[4].alignment = Alignment(horizontal="right", vertical="center")
        row[5].alignment = Alignment(horizontal="right", vertical="center")
        row[6].alignment = Alignment(horizontal="center", vertical="center")

    widths = {"A": 9, "B": 9, "C": 20, "D": 75, "E": 13, "F": 13, "G": 19}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[1].height = 26
    for row_no in range(2, ws.max_row + 1):
        ws.row_dimensions[row_no].height = 42
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_upload_template():
    """대상 목록을 일정한 열 이름으로 작성할 수 있는 빈 엑셀 양식."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "대상목록"
    ws.sheet_view.showGridLines = False

    headers = ["연번", "대상명", "주소", "비고", "위도(선택)", "경도(선택)"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    required_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 첫 입력행은 비워두되 필수 입력칸을 연한 노랑으로 표시한다.
    for row_no in range(2, 102):
        for col_no in range(1, 7):
            cell = ws.cell(row=row_no, column=col_no)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row=row_no, column=2).fill = required_fill
        ws.cell(row=row_no, column=3).fill = required_fill

    widths = {"A": 9, "B": 28, "C": 52, "D": 28, "E": 16, "F": 16}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[1].height = 27
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:F101"

    guide = wb.create_sheet("작성안내")
    guide.sheet_view.showGridLines = False
    guide.append(["항목", "필수 여부", "작성 방법"])
    guide_rows = [
        ["대상명", "필수", "시설명 또는 점검 대상명을 입력합니다."],
        ["주소", "필수", "지오코딩할 도로명주소 또는 지번주소를 입력합니다."],
        ["연번", "선택", "자동으로 표시됩니다. 직접 수정해도 됩니다."],
        ["비고", "선택", "노선 편성에 필요한 일반 참고사항만 입력합니다."],
        ["위도·경도", "선택", "이미 검증한 좌표가 있을 때만 입력합니다. 없으면 비워두세요."],
        ["개인정보", "입력 금지", "성명, 전화번호, 주민등록번호, 검사결과 등은 입력하지 않습니다."],
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in guide.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 14
    guide.column_dimensions["C"].width = 72
    guide.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_printable_qr_html(station, route_results, meta):
    """브라우저에서 열어 A4로 인쇄할 수 있는 노선별 QR 문서를 만든다."""
    cards = []
    title = html.escape(str(meta.get("title") or "순찰노선"))
    period = html.escape(str(meta.get("period") or ""))

    for rr in route_results:
        team_name = html.escape(st.session_state.get(f"team_name_{rr['route_no']}", ""))
        team_members = html.escape(st.session_state.get(f"team_members_{rr['route_no']}", ""))
        qr_blocks = []
        links = kakao_route_links(station, rr["legs"])
        for li, (url, origin, destinations) in enumerate(links, start=1):
            suffix = "" if len(links) == 1 else f" {li}/{len(links)}구간"
            seq = " → ".join([origin["name"]] + [p["name"] for p in destinations])
            qr_b64 = base64.b64encode(make_qr_png(url)).decode("ascii")
            qr_blocks.append(
                f'<section class="qr-block"><h2>노선 {rr["route_no"]}{suffix}</h2>'
                f'<img src="data:image/png;base64,{qr_b64}" alt="노선 QR코드">'
                f'<p class="scan">휴대폰 카메라로 스캔하면 카카오맵 전체 코스가 열립니다.</p>'
                f'<p class="sequence">{html.escape(seq)}</p></section>'
            )
        people = ""
        if team_name or team_members:
            people = f'<p class="people">담당 조 {team_name or "-"}　 조원 {team_members or "-"}</p>'
        cards.append(
            f'<article class="route"><header><div>{title}</div><strong>노선 {rr["route_no"]}</strong>'
            f'<span>{len(rr["stops"])}개소 · {rr["total_km"]:.1f}km · 약 {rr["total_min"]:.0f}분</span>'
            f'</header>{people}{"".join(qr_blocks)}</article>'
        )

    return f'''<!doctype html>
<html lang="ko"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title} QR 인쇄</title>
<style>
@page {{ size: A4; margin: 14mm; }}
* {{ box-sizing: border-box; }}
body {{ margin: 0; color: #111827; font-family: "Malgun Gothic", "Apple SD Gothic Neo", sans-serif; }}
.print-button {{ position: fixed; right: 18px; top: 18px; padding: 12px 18px; border: 0; border-radius: 8px;
  background:#a33a3f; color:white; font-size:16px; font-weight:700; cursor:pointer; }}
.route {{ min-height: 267mm; page-break-after: always; text-align: center; padding: 8mm 5mm; }}
.route:last-child {{ page-break-after: auto; }}
header div {{ font-size: 17px; margin-bottom: 8px; }}
header strong {{ display: block; font-size: 28px; margin-bottom: 6px; }}
header span, .people {{ font-size: 14px; color: #4b5563; }}
.people {{ margin: 10px 0; }}
.qr-block {{ margin-top: 18px; }}
.qr-block h2 {{ font-size: 20px; margin: 0 0 8px; }}
.qr-block img {{ width: 88mm; max-width: 82vw; height: auto; }}
.scan {{ font-size: 14px; font-weight: 700; margin: 4px 0 12px; }}
.sequence {{ font-size: 15px; line-height: 1.7; overflow-wrap: anywhere; border-top: 1px solid #d1d5db;
  padding-top: 12px; margin: 0 auto; max-width: 170mm; }}
.period {{ text-align: center; color: #4b5563; margin: 0 0 8px; }}
@media print {{ .print-button {{ display: none; }} }}
</style></head><body>
<button class="print-button" onclick="window.print()">🖨 인쇄하기</button>
<p class="period">{period}</p>{"".join(cards)}
</body></html>'''.encode("utf-8")


def card_title(step, text):
    st.markdown(
        f'<div class="paseru-card-title"><span class="paseru-step">{step}</span>{text}</div>',
        unsafe_allow_html=True,
    )


def sub_label(text):
    st.markdown(f'<div class="paseru-sub">{text}</div>', unsafe_allow_html=True)


# ---- PWA: 홈 화면에 앱처럼 추가할 수 있도록 매니페스트를 부모 문서에 주입(가능한 환경에서) ----
components.html(
    """
<script>
try {
  const d = window.parent.document;
  if (d && !d.getElementById('paseru-manifest')) {
    const manifest = {
      name: "파세루 오리진 - 순찰노선 설계기",
      short_name: "파세루",
      description: "AI 기반 소방 순찰노선 최적화 서비스",
      start_url: ".", scope: ".", display: "standalone",
      background_color: "#f7f8fa", theme_color: "#a33a3f",
      icons: []
    };
    const link = d.createElement('link');
    link.id = 'paseru-manifest';
    link.rel = 'manifest';
    link.href = 'data:application/manifest+json,' + encodeURIComponent(JSON.stringify(manifest));
    d.head.appendChild(link);
    const meta = d.createElement('meta');
    meta.name = 'apple-mobile-web-app-capable'; meta.content = 'yes';
    d.head.appendChild(meta);
    const theme = d.createElement('meta');
    theme.name = 'theme-color'; theme.content = '#a33a3f';
    d.head.appendChild(theme);
  }
} catch (e) { /* 환경상 주입이 막히면 조용히 무시 */ }
</script>
""",
    height=0,
)

st.markdown(
    """
    <div style="margin:0.2rem 0 1.8rem 0;">
      <div style="margin-bottom:0.38rem;color:#a33a3f;font-size:clamp(0.88rem,1.7vw,1rem);
                  font-weight:650;line-height:1.5;letter-spacing:-0.015em;">
        현장의 주소에서 시작해, 가장 효율적인 소방안전의 길을 만듭니다.
      </div>
      <div style="display:flex;align-items:baseline;gap:0.65rem;flex-wrap:wrap;
                  color:#17263a;line-height:1.2;letter-spacing:-0.035em;">
        <span aria-hidden="true" style="font-size:clamp(1.8rem,4vw,2.45rem);">🚒</span>
        <span style="font-family:'Noto Serif KR',serif;font-size:clamp(2rem,5vw,3rem);
                     font-weight:700;">파세루 오리진</span>
        <span style="color:#667085;font-family:Georgia,serif;font-size:clamp(0.95rem,2vw,1.22rem);
                     font-weight:500;letter-spacing:-0.01em;">FireSafe Route Origin</span>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div style="margin:-0.55rem 0 1.25rem;padding:1.05rem 1.15rem;border:1px solid #cbd3dd;
                border-radius:12px;background:#ffffff;box-shadow:0 2px 10px rgba(23,38,58,.05);">
      <div style="color:#17263a;font-size:1.05rem;font-weight:750;margin-bottom:0.35rem;">
        주소 목록을 실제 도로 기준의 순찰계획으로 바꾸는 소방업무 지원 앱입니다
      </div>
      <div style="color:#344054;font-size:0.94rem;line-height:1.65;">
        여러 대상의 방문 순서를 일일이 지도에서 찾는 대신, 업무 목적과 순찰 조건을 반영해
        출발지부터 복귀지까지의 노선을 자동으로 편성합니다.
      </div>
      <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(165px,1fr));gap:0.55rem;
                  margin-top:0.8rem;color:#263442;font-size:0.88rem;font-weight:600;">
        <div>🛣️ 실제 도로거리·시간 계산</div>
        <div>📍 업무별 방문 순서 편성</div>
        <div>📱 지도·카카오맵·QR 전달</div>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# 공개 주소를 통한 무단 API 사용을 막기 위한 앱 입구 인증
if not st.session_state.get("paseru_authenticated", False):
    with st.container(border=True):
        st.markdown("### 🔐 파세루 오리진 사용자 인증")
        st.caption("이 앱은 승인된 업무 담당자만 이용할 수 있습니다.")
        with st.form("paseru_login_form", clear_on_submit=False):
            entered_password = st.text_input("비밀번호", type="password", placeholder="비밀번호를 입력하세요")
            login_submitted = st.form_submit_button("앱 시작하기", type="primary", use_container_width=True)

        st.markdown(
            '<div style="margin-top:0.45rem;text-align:center;color:#344054;font-size:0.88rem;">'
            '비밀번호를 잊으셨나요? '
            '<a href="mailto:emtmisung@gmail.com?subject=%ED%8C%8C%EC%84%B8%EB%A3%A8%20%EC%98%A4%EB%A6%AC%EC%A7%84%20%EB%B9%84%EB%B0%80%EB%B2%88%ED%98%B8%20%EB%AC%B8%EC%9D%98" '
            'style="color:#a33a3f;font-weight:700;text-decoration:none;">관리자에게 문의</a>'
            '</div>',
            unsafe_allow_html=True,
        )

        if login_submitted:
            if not APP_PASSWORD:
                st.error("관리자가 Streamlit Secrets에 APP_PASSWORD를 먼저 등록해야 합니다.")
            elif hmac.compare_digest(entered_password, APP_PASSWORD):
                st.session_state["paseru_authenticated"] = True
                st.rerun()
            else:
                st.error("비밀번호가 올바르지 않습니다.")
    st.stop()

with st.expander("💡 처음 사용하시나요? 사용 순서와 조건을 설정하는 이유", expanded=False):
    st.markdown(
        """
        **파세루 오리진은 다음 순서로 사용합니다.**

        1. **표지·로그인** — 앱 안내와 개인정보 주의사항을 확인하고 비밀번호로 접속합니다.
        2. **기본정보·대상목록** — 순찰 제목·출발 부서를 입력하고 대상명과 주소만 업로드해 좌표를 확인합니다.
        3. **순찰 세부방법** — 순찰 용도·기간·차량·반복 방식과 출동 여건을 설정합니다.
        4. **노선 생성** — 확정된 좌표와 설정 결과를 확인하고 실제 도로 기준 노선을 계산합니다.
        5. **노선 결과** — 지도·카카오맵·QR·엑셀 결과를 확인하고 내려받습니다.
        """
    )
    st.markdown("**업무별로 조건이 다른 이유**")
    guide_cols = st.columns(2)
    with guide_cols[0]:
        st.markdown(
            "- **지휘관 현장방문:** 하루에 전 현장을 보는 것이 중요해 거리순으로 연결합니다.\n"
            "- **특별경계근무:** 명절·선거·축제의 주요 대상을 하루 1~2회 반복할 수 있습니다.\n"
            "- **계절순찰:** 하루 약 1시간씩 나누고 출동차량의 원거리 이동을 제한합니다."
        )
    with guide_cols[1]:
        st.markdown(
            "- **예방검사:** 검사기한·가능일·팀 수를 계산해 하루 권장량을 정합니다.\n"
            "- **지리조사:** 전체 소화전을 인원과 차량에 균등 배정해 월 10회 안에 점검합니다.\n"
            "- **API 호출 제한:** 반복 계산으로 인한 처리 지연과 지도 API 비용 증가를 막습니다."
        )
    st.caption("자동 생성 노선은 참고안입니다. 현장과 출동 여건을 담당자가 검토한 뒤 최종 노선을 결정하세요.")

if not has_keys():
    st.error(
        "NCP(네이버클라우드플랫폼) Client ID/Secret이 설정되지 않았습니다. "
        "`.streamlit/secrets.toml` 또는 Streamlit Cloud의 Secrets 설정에 "
        "NCP_CLIENT_ID / NCP_CLIENT_SECRET 값을 등록해주세요."
    )


def next_tab_button(label, target_index):
    """현재 입력을 유지한 채 다음 Streamlit 탭으로 이동한다."""
    target_names = ["② 기본정보 · 대상목록", "③ 순찰 세부방법", "④ 노선 생성", "⑤ 노선 결과"]
    target_name = target_names[target_index]
    components.html(
        f"""
        <style>
          .paseru-next {{
            width:33.333%; min-width:240px; padding:11px 12px 11px 22px;
            border:2px solid #207447; border-radius:999px;
            background:linear-gradient(135deg,#238553,#17663e); color:#fff;
            font-family:Arial,'Noto Sans KR',sans-serif; font-size:15px; font-weight:800;
            cursor:pointer; box-shadow:0 8px 20px -9px rgba(23,102,62,.8);
            display:flex; align-items:center; justify-content:space-between; gap:12px;
            transition:transform .16s ease,box-shadow .16s ease,filter .16s ease;
          }}
          .paseru-next:hover {{
            transform:translateY(-2px); filter:brightness(1.06);
            box-shadow:0 12px 24px -10px rgba(23,102,62,.9);
          }}
          .paseru-next .arrow {{
            width:34px; height:34px; flex:0 0 34px; border-radius:50%;
            display:flex; align-items:center; justify-content:center;
            background:#fff; color:#17663e; font-size:20px; font-weight:900;
          }}
        </style>
        <button onclick="goNext()"
          class="paseru-next" title="입력을 저장하고 다음 단계로 이동합니다">
          <span>다음 단계로 이동 · {html.escape(label.replace('다음 · ', ''))}</span>
          <span class="arrow">→</span>
        </button>
        <script>
          function goNext() {{
            const doc = window.parent.document;
            const candidates = Array.from(doc.querySelectorAll(
              'button[role="tab"], [data-baseweb="tab-list"] button, button[data-baseweb="tab"]'
            ));
            const wanted = '{target_name}';
            const byName = candidates.find(el => (el.innerText || el.textContent || '').trim().includes(wanted));
            const target = byName || candidates[{target_index}];
            if (target) {{
              target.scrollIntoView({{behavior:'smooth', block:'start'}});
              target.dispatchEvent(new MouseEvent('click', {{bubbles:true, cancelable:true, view:window.parent}}));
              target.click();
            }} else {{
              alert('다음 탭을 찾지 못했습니다. 상단의 {target_name} 탭을 눌러주세요.');
            }}
          }}
        </script>
        """,
        height=76,
    )

# ----------------------------------------------------------------------------
page_basic, page_details, page_build, page_results = st.tabs([
    "② 기본정보 · 대상목록",
    "③ 순찰 세부방법",
    "④ 노선 생성",
    "⑤ 노선 결과",
])

with page_basic:
    # 1 · 기본 정보 / 대상 목록
    # ----------------------------------------------------------------------------
    with st.container(border=True):
        card_title(2, "기본 정보 · 대상 목록")
        patrol_title = st.text_input("순찰 제목", value="예시) 소방안전 순찰노선 - 성주군 일원")
        c1, c2 = st.columns([1, 1.8])
        with c1:
            station_name = st.text_input("출발 부서(소방서·센터) 이름", value="성주소방서")
        with c2:
            station_address = st.text_input("출발 부서 주소", value="경상북도 성주군 성주읍 주산로 193")
        route_prefix = station_name

        st.markdown("**대상 목록 업로드**")
        st.markdown(
            """
            <div style="margin:0.25rem 0 1rem;padding:1rem 1.1rem;border:1px solid #d7a54a;
                        border-left:5px solid #b7791f;border-radius:10px;background:#fff7e8;
                        color:#5f3d0c;line-height:1.6;">
              <div style="font-size:1.08rem;font-weight:750;margin-bottom:0.25rem;color:#744b0f;">
                ⚠️ 개인정보가 포함된 파일은 업로드하지 마세요
              </div>
              <div style="font-size:0.96rem;font-weight:600;color:#5f3d0c;">
                이 앱은 공개 앱입니다. 대상명과 주소만 입력하고, 개인 성명·담당자 실명·전화번호 등
                개인정보와 민감정보는 파일에 포함하지 마세요.
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        template_col, template_note_col = st.columns([1, 2])
        with template_col:
            st.download_button(
                "📥 대상 목록 빈 양식(xlsx)", data=build_upload_template(),
                file_name="파세루_대상목록_빈양식.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with template_note_col:
            st.caption("양식을 내려받아 노란색 `대상명·주소` 칸만 작성하세요. 개인정보와 민감정보는 입력하지 마세요.")
        uploaded = st.file_uploader("대상 목록 파일", type=["csv", "xlsx", "xls", "hwpx"],
                                    label_visibility="collapsed")
        use_sample = st.checkbox("🧪 기능 확인용 예시 20건 불러오기 (성주군 주요 대상)",
                                 value=uploaded is None)

    df = None
    if uploaded is not None:
        name_lower = uploaded.name.lower()
        if name_lower.endswith(".csv"):
            df = pd.read_csv(uploaded)
        elif name_lower.endswith(".hwpx"):
            df = parse_hwpx(uploaded.getvalue())
            if df is None:
                st.error("hwpx 파일에서 표나 목록을 찾지 못했습니다. 표 형식인지 확인해주세요.")
        else:
            df = pd.read_excel(uploaded)
    elif use_sample:
        df = pd.read_excel(SAMPLE_XLSX)


    @st.cache_resource
    def coordinate_executor():
        return ThreadPoolExecutor(max_workers=2, thread_name_prefix="paseru-coordinates")


    def search_coordinates_in_background(records, name_key, address_key, lat_key=None, lng_key=None):
        rows = []
        for record in records:
            nm, ad = str(record.get(name_key, "")), str(record.get(address_key, ""))
            file_lat = file_lng = None
            if lat_key and lng_key:
                try:
                    file_lat, file_lng = float(record.get(lat_key)), float(record.get(lng_key))
                    if math.isnan(file_lat) or math.isnan(file_lng):
                        file_lat = file_lng = None
                except (TypeError, ValueError):
                    file_lat = file_lng = None
            if file_lat is not None:
                rows.append({"대상명": nm, "주소": ad, "위도": file_lat, "경도": file_lng,
                             "상태": "파일 좌표", "비고": "파일에 있던 좌표를 사용"})
                continue
            lat, lng, used_q, used_why, tried = geocode_with_fallback(ad, nm)
            if lat is None:
                rows.append({"대상명": nm, "주소": ad, "위도": None, "경도": None,
                             "상태": "❌ 실패", "비고": "시도: " + " / ".join(tried)})
            else:
                rows.append({"대상명": nm, "주소": ad, "위도": lat, "경도": lng,
                             "상태": "✅ 확인" if used_why == "원본 주소" else "🔧 주소 보정 후 확인",
                             "비고": "" if used_why == "원본 주소" else f"{used_why} → {used_q}"})
        return pd.DataFrame(rows)


    if df is not None and len(df):
        pre_cols = list(df.columns)
        pre_name_idx = next((i for i, c in enumerate(pre_cols) if str(c) not in ("연번",) and
                             ("주소지" in str(c) or "이름" in str(c) or "명" in str(c))),
                            1 if len(pre_cols) > 1 else 0)
        pre_addr_idx = next((i for i, c in enumerate(pre_cols) if "정제" in str(c)),
                            next((i for i, c in enumerate(pre_cols) if "주소" in str(c) and i != pre_name_idx),
                                 min(3, len(pre_cols) - 1)))
        pre_lat = next((c for c in pre_cols if "위도" in str(c) or str(c).lower() == "lat"), None)
        pre_lng = next((c for c in pre_cols if "경도" in str(c) or str(c).lower() in ("lng", "lon")), None)
        coord_signature = tuple((str(row[pre_cols[pre_name_idx]]), str(row[pre_cols[pre_addr_idx]]))
                                for _, row in df.iterrows())
        if st.session_state.get("coord_signature") != coord_signature:
            st.session_state["coord_signature"] = coord_signature
            st.session_state.pop("coords_df", None)
            st.session_state.pop("coord_future", None)

        with st.container(border=True):
            st.markdown("### 🔎 대상 좌표 우선 확인")
            coord_future = st.session_state.get("coord_future")
            saved_early = st.session_state.get("coords_df")
            if coord_future is None and saved_early is None:
                coord_button_col, _, _ = st.columns(3)
                with coord_button_col:
                    if st.button("🔴 좌표 검색 시작", type="primary", use_container_width=True,
                                 disabled=not has_keys()):
                        st.session_state["coord_future"] = coordinate_executor().submit(
                            search_coordinates_in_background, df.to_dict("records"),
                            pre_cols[pre_name_idx], pre_cols[pre_addr_idx], pre_lat, pre_lng,
                        )
                        st.rerun()
            elif coord_future is not None and not coord_future.done():
                @st.fragment(run_every="1s")
                def poll_coordinate_search():
                    running = st.session_state.get("coord_future")
                    if running is not None and running.done():
                        try:
                            st.session_state["coords_df"] = running.result()
                            st.session_state.pop("coord_future", None)
                            st.rerun()
                        except Exception as exc:
                            st.session_state.pop("coord_future", None)
                            st.error(f"⚠️ 좌표 검색 중 오류가 발생했습니다: {type(exc).__name__}")
                    else:
                        st.info("🔍 좌표를 검색하고 있습니다…  ◀︎ 🔎 ▶︎  아래 순찰 용도와 일정을 계속 입력하세요.")
                poll_coordinate_search()
            elif coord_future is not None:
                try:
                    st.session_state["coords_df"] = coord_future.result()
                    st.session_state.pop("coord_future", None)
                    st.rerun()
                except Exception as exc:
                    st.session_state.pop("coord_future", None)
                    st.error(f"⚠️ 좌표 검색 중 오류가 발생했습니다: {type(exc).__name__}")
            else:
                fail_early = int(saved_early["위도"].isna().sum())
                if fail_early:
                    st.warning(f"⚠️ 좌표 검색 완료 · 성공 {len(saved_early)-fail_early}건 · 실패 {fail_early}건")
                else:
                    coord_result_col, _, _ = st.columns(3)
                    with coord_result_col:
                        with st.expander(f"🙂 좌표 검색 완료 · {len(saved_early)}건 결과 보기", expanded=False):
                            st.dataframe(saved_early, use_container_width=True, hide_index=True)

    st.write("")

    basic_ready = bool(
        patrol_title.strip() and station_name.strip() and station_address.strip()
        and df is not None and len(df) and st.session_state.get("coords_df") is not None
        and st.session_state.get("coord_future") is None
    )
    if basic_ready:
        next_tab_button("다음 · ③ 순찰 세부방법", 1)
    else:
        st.caption("제목·출발부서·대상목록을 입력하고 좌표 검색이 끝나면 다음 버튼이 나타납니다.")

    # ----------------------------------------------------------------------------

with page_details:
    # 2 · 순찰 방법과 세부 일정
    # ----------------------------------------------------------------------------
    PURPOSE_OPTIONS = [
        "① 지휘관 현장방문", "② 특별경계근무용", "③ 계절순찰", "④ 예방검사", "⑤ 지리조사(센터용)",
    ]
    PURPOSE_HINT = {
        "① 지휘관 현장방문": "풍수해·산불·재난지역 등 모든 현장을 하루에 실제 도로거리순으로 방문합니다.",
        "② 특별경계근무용": "명절·선거·축제 등 특별경계근무 — 휴무 공장과 터미널·역·공항·행사장 등 주요 대상을 하루 1~2회 반복 순찰합니다.",
        "③ 계절순찰": "정해진 기간 동안 수행자·차량·편도 제한·1회 최대시간을 반영해 반복형 또는 전 대상 순환형 노선을 만듭니다.",
        "④ 예방검사": "숙박업소 등 점검 순찰.",
        "⑤ 지리조사(센터용)": "소화전 등 팀별 순회 — 팀 수·목표시간 기준으로 노선수를 자동 산출합니다.",
    }

    with st.container(border=True):
        card_title(3, "순찰 방법 · 세부 조건")
        purpose_label = st.pills("노선 용도", PURPOSE_OPTIONS, default=None,
                                 label_visibility="collapsed")
        if not purpose_label:
            st.info("먼저 순찰방법을 하나 선택하면 대상 업로드와 세부 설정이 나타납니다.")
            st.stop()
        if hasattr(st, "popover"):
            with st.popover("❔ 선택한 순찰방법 설명 보기"):
                st.write(PURPOSE_HINT.get(purpose_label, ""))
                st.caption("순찰방법을 바꾸면 해당 업무에 필요한 조건만 자동으로 표시됩니다.")
        else:
            with st.expander("❔ 선택한 순찰방법 설명 보기", expanded=False):
                st.write(PURPOSE_HINT.get(purpose_label, ""))
        purpose = {
            "① 지휘관 현장방문": "other", "② 특별경계근무용": "guard",
            "③ 계절순찰": "season", "④ 예방검사": "inspect",
            "⑤ 지리조사(센터용)": "hydrant",
        }.get(purpose_label, "other")

        guard_repeat_label = None
        guard_rounds = None
        hydrant_members = []
        hydrant_member_count = 0
        hydrant_vehicle_count = 0
        hydrant_workdays = 10
        hydrant_target_min = 90
        hydrant_max_min = 120
        hydrant_inspection_min = 5
        season_scope = "관할 전체 대상 균등 순환"
        season_actor = "소방공무원"
        season_vehicle = "소방차"
        season_limit_basis = "편도시간(분)"
        season_oneway_limit = 30
        season_strategy = "전 대상 균등 순환"
        season_delegate = "의용소방대 순찰 권장"
        commander_route_mode = "전체 대상을 하나의 노선으로 연결"
        commander_vehicle_count = 1

        if purpose == "guard":
            gc1, gc2 = st.columns([1.6, 1])
            with gc1:
                sub_label("반복 방식")
                guard_repeat_label = st.pills("반복 방식", ["매일 같은 코스 반복", "매일 다른 코스 순환"],
                                              default="매일 같은 코스 반복", label_visibility="collapsed")
            with gc2:
                if guard_repeat_label == "매일 같은 코스 반복":
                    sub_label("하루 반복 횟수")
                    guard_rounds = st.pills("하루 반복 횟수", ["1회", "2회", "3회"], default="1회",
                                            label_visibility="collapsed")
        elif purpose == "season":
            st.caption("업로드한 관할 전체 대상을 기간 동안 서로 다른 코스로 균등하게 순환합니다.")
            sc1, sc2 = st.columns(2)
            with sc1:
                season_actor = st.pills(
                    "누가 순찰합니까?", ["소방공무원", "의용소방대"], default="소방공무원"
                ) or "소방공무원"
            with sc2:
                season_vehicle = st.pills(
                    "어떤 차량을 이용합니까?", ["소방차", "구급차", "행정차", "일반차"], default="소방차"
                ) or "소방차"
            sc3, sc4 = st.columns(2)
            with sc3:
                season_limit_basis = st.pills(
                    "센터 기준 편도 제한", ["편도시간(분)", "편도거리(km)"], default="편도시간(분)"
                ) or "편도시간(분)"
            with sc4:
                season_oneway_limit = st.number_input(
                    "편도 제한값",
                    min_value=1.0,
                    max_value=120.0,
                    value=30.0 if season_limit_basis == "편도시간(분)" else 20.0,
                    step=5.0 if season_limit_basis == "편도시간(분)" else 1.0,
                    help="편도 30분은 왕복 이동만 약 1시간인 거리입니다. 기준을 넘으면 의용소방대 순찰을 권장합니다.",
                )
            season_delegate = ("의용소방대 순찰 권장" if season_actor == "소방공무원"
                               else "장거리 별도 순찰 검토")
        elif purpose == "hydrant":
            hc1, hc2 = st.columns(2)
            with hc1:
                hydrant_member_count = st.number_input("지리조사 인원 수", min_value=1, max_value=30, value=2)
            with hc2:
                hydrant_vehicle_count = st.number_input("운행 차량 수", min_value=1, max_value=15, value=1)
            st.caption("전체 소화전을 인원수로 균등 배정하고, 같은 차량 팀원의 담당 구역은 서로 가깝게 묶습니다.")
            with st.expander("차량별 팀원 편성", expanded=True):
                vehicle_options = list(range(1, int(hydrant_vehicle_count) + 1))
                for member_index in range(int(hydrant_member_count)):
                    mc1, mc2 = st.columns([1.6, 1])
                    default_vehicle = vehicle_options[member_index % len(vehicle_options)]
                    with mc1:
                        member_name = st.text_input(
                            f"팀원 {member_index + 1}",
                            value=f"대원{member_index + 1}",
                            key=f"hydrant_member_name_{member_index}",
                            help="공개 앱에서는 실명 대신 대원1, 대원2 같은 호출명을 권장합니다.",
                        ).strip() or f"대원{member_index + 1}"
                    with mc2:
                        vehicle_no = st.selectbox(
                            f"팀원 {member_index + 1} 차량",
                            vehicle_options,
                            index=vehicle_options.index(default_vehicle),
                            format_func=lambda value: f"{value}호차",
                            key=f"hydrant_member_vehicle_{member_index}",
                        )
                    hydrant_members.append({
                        "name": member_name,
                        "vehicle_no": int(vehicle_no),
                        "order": member_index,
                    })
        elif purpose == "other":
            visit_purpose = st.pills(
                "방문 목적", ["풍수해 현장", "산불 현장", "재난 현장", "기타"], default="재난 현장"
            ) or "재난 현장"
            commander_route_mode = st.pills(
                "노선 구성",
                ["전체 대상을 하나의 노선으로 연결", "여러 차량으로 균등 분할"],
                default="전체 대상을 하나의 노선으로 연결",
            ) or "전체 대상을 하나의 노선으로 연결"
            if commander_route_mode == "여러 차량으로 균등 분할":
                commander_vehicle_count = st.number_input("방문 차량 수", min_value=2, max_value=20, value=2)
            commander_summary = (
                f"① 지휘관 현장방문 · {visit_purpose} · {commander_route_mode} · "
                "시간 제한 없이 실제 도로거리상 가까운 순서로 연결"
            )
            st.caption("입력한 조건은 좌표 검색 결과와 결합한 뒤 ⑤ 노선 생성 단계의 ‘설정 결과’에서 확인합니다.")

    st.write("")

    # ----------------------------------------------------------------------------
    # 2 · 순찰 일정
    # ----------------------------------------------------------------------------
    if "period_start" not in st.session_state:
        st.session_state["period_start"] = date(2026, 9, 23)
        st.session_state["period_start_time"] = dtime(18, 0)
        st.session_state["period_end"] = date(2026, 9, 28)
        st.session_state["period_end_time"] = dtime(9, 0)

    with st.container(border=True):
        if purpose == "inspect":
            card_title(3, "예방검사 일정")
            st.caption("대상 파일에는 대상명과 주소만 준비하면 됩니다. 공통 검사 조건은 여기에서 한 번만 설정합니다.")

            ic1, ic2 = st.columns(2)
            with ic1:
                period_start = st.date_input("검사 시작일", key="period_start")
            with ic2:
                period_end = st.date_input("검사 완료기한", key="period_end")

            weekday_names = ["월", "화", "수", "목", "금", "토", "일"]
            inspect_weekdays = st.multiselect(
                "검사 가능 요일",
                weekday_names,
                default=["월", "화", "수", "목", "금"],
                help="실제로 예방검사를 실시할 요일만 선택하세요.",
            )

            ic3, ic4, ic5 = st.columns(3)
            with ic3:
                inspect_teams = st.number_input("검사팀 수", min_value=1, max_value=30, value=1)
            with ic4:
                inspect_daily_hours = st.number_input(
                    "팀당 하루 검사 가능시간", min_value=1.0, max_value=12.0, value=6.0, step=0.5,
                )
            with ic5:
                inspect_minutes = st.number_input(
                    "대상당 평균 검사시간(분)", min_value=5, max_value=480, value=40, step=5,
                )

            excluded_text = st.text_input(
                "검사 제외일(선택)",
                placeholder="예) 2026-09-21, 2026-10-03",
                help="공휴일·훈련일 등 검사하지 않는 날짜를 쉼표로 구분해 입력하세요.",
            )
            inspect_excluded_dates = set()
            invalid_excluded_dates = []
            for value in [v.strip() for v in excluded_text.split(",") if v.strip()]:
                try:
                    inspect_excluded_dates.add(datetime.strptime(value, "%Y-%m-%d").date())
                except ValueError:
                    invalid_excluded_dates.append(value)
            if invalid_excluded_dates:
                st.warning("제외일은 YYYY-MM-DD 형식으로 입력해주세요: " + ", ".join(invalid_excluded_dates))

            start_dt = datetime.combine(period_start, dtime(9, 0))
            end_dt = datetime.combine(period_end, dtime(18, 0))
            selected_weekdays = {i for i, name in enumerate(weekday_names) if name in inspect_weekdays}
            inspect_dates = []
            if period_end < period_start:
                st.warning("⚠ 검사 완료기한이 시작일보다 빠릅니다. 기간을 확인해주세요.")
            elif not selected_weekdays:
                st.warning("⚠ 검사 가능 요일을 하나 이상 선택해주세요.")
            else:
                current_date = period_start
                while current_date <= period_end:
                    if current_date.weekday() in selected_weekdays and current_date not in inspect_excluded_dates:
                        inspect_dates.append(current_date)
                    current_date += timedelta(days=1)

            period_days = max(1, len(inspect_dates))
            st.caption(
                f"실제 검사 가능일 {len(inspect_dates)}일 · 전체 가용 팀 일수 "
                f"{len(inspect_dates) * int(inspect_teams)}팀 일"
            )
            vehicle = st.selectbox("검사 차량", ["소방차", "구급차", "행정차", "개인차"], index=2)
        elif purpose == "season":
            card_title(3, "계절순찰 일정")
            dc1, dc2 = st.columns(2)
            with dc1:
                period_start = st.date_input("순찰 시작일", key="period_start")
            with dc2:
                period_end = st.date_input("순찰 종료일", key="period_end")
            season_time_choice = st.pills(
                "1회 순찰시간 선택",
                ["30분", "60분", "직접 입력"],
                default="60분",
                help="출발·복귀 이동과 대상별 현장 확인시간을 모두 포함한 시간입니다.",
            ) or "60분"
            sc_time1, sc_time2 = st.columns(2)
            with sc_time1:
                if season_time_choice == "직접 입력":
                    season_target_min = st.number_input(
                        "직접 입력(분)", min_value=20, max_value=360, value=90, step=10,
                    )
                else:
                    season_target_min = int(season_time_choice.replace("분", ""))
                    st.metric("적용할 1회 최대시간", f"{season_target_min}분")
            with sc_time2:
                season_stop_min = st.number_input(
                    "대상당 현장 확인시간(분)", min_value=0, max_value=30, value=2,
                )
            start_dt = datetime.combine(period_start, dtime(9, 0))
            end_dt = datetime.combine(period_end, dtime(18, 0))
            period_days = max(1, (period_end - period_start).days + 1)
            vehicle = season_vehicle
            inspect_weekdays = []
            inspect_teams = 1
            inspect_daily_hours = 6.0
            inspect_minutes = 40
            inspect_dates = []
            st.caption("입력한 조건은 좌표 검색 결과와 결합한 뒤 ⑤ 노선 생성 단계의 ‘설정 결과’에서 확인합니다.")
        elif purpose == "hydrant":
            card_title(3, "월간 지리조사 설정")
            st.caption("당비비 근무 기준으로 한 달 10번의 당번일 안에 전체 소화전을 점검하도록 노선을 나눕니다.")
            hc1, hc2 = st.columns(2)
            with hc1:
                survey_year = st.number_input("조사 연도", min_value=2024, max_value=2100, value=2026)
            with hc2:
                survey_month = st.selectbox("조사 월", list(range(1, 13)), index=8,
                                            format_func=lambda value: f"{value}월")

            hc3, hc4, hc5, hc6 = st.columns(4)
            with hc3:
                hydrant_workdays = st.number_input("월 당번 근무일", min_value=1, max_value=31, value=10)
            with hc4:
                hydrant_target_min = st.number_input(
                    "노선 기본 목표시간(분)", min_value=60, max_value=240, value=90, step=10,
                )
            with hc5:
                hydrant_max_min = st.number_input(
                    "노선 최대 허용시간(분)", min_value=60, max_value=360, value=120, step=10,
                )
            with hc6:
                hydrant_inspection_min = st.number_input(
                    "소화전 1개 조사시간(분)", min_value=0, max_value=60, value=5, step=1,
                )

            if hydrant_max_min < hydrant_target_min:
                st.warning("최대 허용시간은 기본 목표시간보다 길게 설정해주세요.")
                hydrant_max_min = hydrant_target_min
            last_day = calendar.monthrange(int(survey_year), int(survey_month))[1]
            period_start = date(int(survey_year), int(survey_month), 1)
            period_end = date(int(survey_year), int(survey_month), last_day)
            start_dt = datetime.combine(period_start, dtime(0, 0))
            end_dt = datetime.combine(period_end, dtime(23, 59))
            period_days = int(hydrant_workdays)
            vehicle = f"소방차 {int(hydrant_vehicle_count)}대"
            inspect_weekdays = []
            inspect_teams = 1
            inspect_daily_hours = 6.0
            inspect_minutes = 40
            inspect_dates = []
            st.caption(
                f"기본 {int(hydrant_target_min)}분 이내로 편성하고, 차량별 노선이 "
                f"{int(hydrant_workdays)}개를 넘으면 시간을 늘리도록 안내합니다."
            )
        elif purpose == "other":
            card_title(3, "지휘관 현장방문 일정")
            visit_date = st.date_input("현장방문일", value=date.today(), key="commander_visit_date")
            commander_vehicle = st.selectbox("방문 차량", ["지휘차", "행정차", "소방차", "기타"])
            period_start = period_end = visit_date
            start_dt = datetime.combine(visit_date, dtime(9, 0))
            end_dt = datetime.combine(visit_date, dtime(18, 0))
            period_days = 1
            vehicle = commander_vehicle
            inspect_weekdays = []
            inspect_teams = 1
            inspect_daily_hours = 6.0
            inspect_minutes = 40
            inspect_dates = []
            st.caption("시간 제한 없이 선택한 모든 현장을 하루 동안 실제 도로거리순으로 방문합니다.")
        else:
            card_title(3, "순찰 기간 · 순찰 차량")
            inspect_weekdays = []
            inspect_teams = 1
            inspect_daily_hours = 6.0
            inspect_minutes = 40
            inspect_dates = []

            dc1, dc2, dc3, dc4 = st.columns(4)
            with dc1:
                period_start = st.date_input("시작일", key="period_start")
            with dc2:
                period_start_time = st.time_input("시작 시각", key="period_start_time")
            with dc3:
                period_end = st.date_input("종료일", key="period_end")
            with dc4:
                period_end_time = st.time_input("종료 시각", key="period_end_time")

            start_dt = datetime.combine(period_start, period_start_time)
            end_dt = datetime.combine(period_end, period_end_time)
            if end_dt <= start_dt:
                st.warning("⚠ 종료 일시가 시작 일시보다 빠릅니다. 기간을 확인해주세요.")
                period_days = 1
            else:
                period_days = max(1, math.ceil((end_dt - start_dt).total_seconds() / 86400))
                st.caption(f"총 {period_days}일간")

            vehicle = st.selectbox("순찰 차량", ["소방차", "구급차", "행정차", "개인차"], index=0)

    st.write("")

    # ----------------------------------------------------------------------------
    # 4 · 노선 조건 설정
    # ----------------------------------------------------------------------------
    with st.expander("💡 ④ 상세 노선 조건 보기", expanded=False):
        card_title(3, "노선 조건 설정")

        if purpose == "season":
            mode = "target_time"
            target_min = int(season_target_min)
            target_min_low = None
            target_min_high = target_min
            seg_max_km = seg_max_min = None
            max_per_route = 100
            max_routes_cap = 0
            basis_label = "소요시간 기준"
            basis = "time"
            st.markdown("**계절순찰 자동 편성 기준**")
            st.caption(
                f"1회 최대 {target_min}분 · 관할 전체 대상을 코스별로 균등 순환합니다. "
                f"{season_vehicle}의 센터 기준 {season_limit_basis} 제한값은 {season_oneway_limit:g}이며, "
                f"이를 넘는 대상은 {season_delegate} 대상으로 분리해 안내합니다."
            )
        elif purpose == "hydrant":
            mode = "target_time"
            target_min = int(hydrant_target_min)
            target_min_low = 60
            target_min_high = int(hydrant_target_min)
            seg_max_km = seg_max_min = None
            max_per_route = 100
            max_routes_cap = 0
            basis_label = "소요시간 기준"
            basis = "time"
            st.markdown("**지리조사 자동 편성 기준**")
            st.caption(
                f"개인별 개수를 균등하게 배정한 뒤 같은 차량 팀원의 인접 구역을 묶고, "
                f"센터 출발·복귀 포함 {target_min}분 이내 노선으로 나눕니다."
            )
        elif purpose == "other":
            mode = "fixed"
            target_min = None
            target_min_low = target_min_high = None
            seg_max_km = seg_max_min = None
            max_per_route = 100
            max_routes_cap = (int(commander_vehicle_count)
                              if commander_route_mode == "여러 차량으로 균등 분할" else 1)
            basis_label = "거리 기준"
            basis = "distance"
            st.markdown("**지휘관 현장방문 자동 편성 기준**")
            st.caption(
                "시간 제한 없이 전 대상을 가까운 순서로 연결합니다. "
                + (f"차량 {int(commander_vehicle_count)}대에 균등하게 나눕니다."
                   if commander_route_mode == "여러 차량으로 균등 분할"
                   else "기본적으로 하루 한 개 노선으로 편성합니다.")
            )
        else:
            sub_label("가. 기준 방식")
            mode_label = st.pills("기준 방식",
                                  ["노선 수·구간 수 지정", "구간별 제한", "노선 전체 목표시간"],
                                  default="노선 수·구간 수 지정", label_visibility="collapsed")
            if not mode_label:
                mode_label = "노선 수·구간 수 지정"
            mode = {"노선 수·구간 수 지정": "fixed", "구간별 제한": "segment",
                    "노선 전체 목표시간": "target_time"}[mode_label]

            cc1, cc2 = st.columns(2)
            with cc1:
                max_per_route = st.number_input("노선 내 구간 수(방문지 수)", min_value=1, max_value=30, value=5)
            with cc2:
                max_routes_cap = st.number_input("총 노선 수 상한(0 = 전수 자동배분)", min_value=0, value=6)

            if mode == "fixed":
                st.caption(f"노선당 {max_per_route}개소씩 최대 {max_routes_cap or '제한 없이'}개 노선으로 나눕니다. "
                           "거리·시간 제한 없이 개수대로 나눈 뒤, 아래 목표시간을 넘는 노선은 표시해 드립니다.")
                target_min = st.number_input("참고용 목표 왕복시간(분) — 초과 노선을 표시만 합니다",
                                             min_value=10, value=30)
                target_min_low = target_min_high = None
                seg_max_km = seg_max_min = None
            elif mode == "segment":
                sc1, sc2 = st.columns(2)
                with sc1:
                    seg_max_km = st.number_input("구간당 최대 거리(km)", min_value=1.0, value=7.0, step=0.5)
                with sc2:
                    seg_max_min = st.number_input("구간당 최대 시간(분)", min_value=1, value=10)
                target_min = target_min_low = target_min_high = None
            else:
                sub_label("나. 순찰 소요시간(왕복 목표시간)")
                quick_min = st.pills("목표 시간", ["30분", "1시간", "2시간", "직접입력"],
                                     default="1시간", label_visibility="collapsed")
                if quick_min == "30분":
                    target_min = 30
                elif quick_min == "2시간":
                    target_min = 120
                elif quick_min == "직접입력":
                    target_min = st.number_input("목표 왕복시간(분)", min_value=10, value=90)
                else:
                    target_min = 60
                allow_range = st.slider("허용 범위(분, ±)", 0, 60, 15)
                target_min_low = target_min - allow_range
                target_min_high = target_min + allow_range
                seg_max_km = seg_max_min = None

        if purpose not in ("hydrant", "season", "other"):
            sub_label("다. 노선 생성 기준")
            basis_label = st.pills("노선 생성 기준", ["거리 기준", "소요시간 기준"],
                                   default="거리 기준", label_visibility="collapsed")
            if not basis_label:
                basis_label = "거리 기준"
            basis = "time" if basis_label == "소요시간 기준" else "distance"
            st.caption("거리 기준: 이동 거리(km)가 가장 짧은 순서로 연결 / 소요시간 기준: 이동 시간(분)이 가장 짧은 순서로 연결")

            sub_label("라. 장거리 분리 기준")
            long_threshold = st.number_input("소방서 실제 도로거리(km) 초과 시 별도 표시", min_value=1.0, value=15.0)
        elif purpose == "hydrant":
            long_threshold = 99999.0
            st.caption("월간 전수조사이므로 장거리 소화전도 분리하지 않고 반드시 차량·팀원에게 배정합니다.")
        elif purpose == "season":
            long_threshold = 99999.0
            st.caption(
                f"직선거리가 아닌 실제 도로의 {season_limit_basis} 기준으로 판단합니다. "
                f"제한값 {season_oneway_limit:g} 초과 대상은 별도 검토 대상으로 표시합니다."
            )
        else:
            long_threshold = 99999.0
            st.caption("지휘관 현장방문은 원거리 대상을 제외하지 않고 전 대상을 거리순으로 편성합니다.")

        with st.expander("⚙️ 계산 과정·API 호출 설정 보기", expanded=False):
            st.caption("기본값 그대로 사용해도 됩니다. 비용이나 계산 정밀도를 직접 조정할 때만 변경하세요.")
            save_calls = st.checkbox(
                "가까운 후보만 실제 도로거리로 확인하여 호출 절약 (권장)", value=True,
                help="직선거리는 후보를 좁히는 데만 사용하며, 최종 노선은 실제 도로거리로 계산합니다. "
                     "끄면 남은 모든 대상을 실제 도로거리로 확인해 호출 횟수가 크게 늘어납니다.",
            )
            if save_calls:
                candidate_k = st.slider("실제 도로거리로 확인할 후보 수", 3, 12, 5,
                                        help="숫자가 클수록 비교 후보는 많아지지만 API 호출도 늘어납니다.")
            else:
                candidate_k = 0

            max_calls = st.number_input(
                "최대 API 호출 수", min_value=50, max_value=100000,
                value=1500, step=100,
                help="설정한 횟수에 도달하면 비용 보호를 위해 추가 호출을 중단합니다.",
            )

    st.write("")
    next_tab_button("다음 · ④ 노선 생성", 2)


with page_build:
    # 출발지(소방서·센터) 자신이 순찰 대상 목록에 섞여 있으면 제외한다.
    # (업로드 파일 첫 줄에 소방서를 넣어두는 경우가 많아, 그대로 두면 소방서가 경유지로 잡힌다)
    excluded_station_rows = 0
    if df is not None and len(df):
        def _norm(v):
            return re.sub(r"\s+", "", str(v)) if v is not None else ""

        st_name_n, st_addr_n = _norm(station_name), _norm(station_address)
        mask_keep = []
        for _, r in df.iterrows():
            vals = [_norm(v) for v in r.values]
            is_station = any(v and (v == st_name_n or v == st_addr_n) for v in vals)
            mask_keep.append(not is_station)
        excluded_station_rows = len(df) - sum(mask_keep)
        if excluded_station_rows:
            df = df[pd.Series(mask_keep, index=df.index)].reset_index(drop=True)

    if st.session_state.get("route_results"):
        next_tab_button("다음 · ⑤ 노선 결과", 3)
    else:
        st.caption("노선 생성이 완료되면 결과 화면으로 이동하는 다음 버튼이 나타납니다.")

    # ----------------------------------------------------------------------------
    # 5 · 미리보기 · 노선 생성
    # ----------------------------------------------------------------------------
    if df is not None and len(df):
        st.write("")
        with st.container(border=True):
            card_title(4, "노선 생성")
            data_check_col, _ = st.columns([1, 1])
            with data_check_col:
                with st.expander(f"🔎 업로드 데이터 확인 · 총 {len(df)}건", expanded=False):
                    st.dataframe(df, use_container_width=True, hide_index=True)
            if excluded_station_rows:
                st.info(f"ℹ️ 목록에 있던 출발지({station_name}) {excluded_station_rows}건은 "
                        "순찰 대상이 아니라 출발·복귀 지점이므로 자동으로 제외했습니다.")

            cols = list(df.columns)
            # 이름 컬럼 기본값: "연번" 같은 숫자 컬럼이 아니라 실제 명칭이 담긴 컬럼을 우선 선택
            name_col_guess_idx = next(
                (i for i, c in enumerate(cols)
                 if str(c) not in ("연번",) and ("주소지" in str(c) or "이름" in str(c) or "명" in str(c))),
                None,
            )
            if name_col_guess_idx is None:
                name_col_guess_idx = 1 if len(cols) > 1 else 0
            # 주소 컬럼 기본값: "정제_주소"처럼 지오코딩에 바로 쓸 수 있는 컬럼을 최우선으로
            addr_col_guess_idx = next(
                (i for i, c in enumerate(cols) if "정제" in str(c)),
                next((i for i, c in enumerate(cols)
                      if "주소" in str(c) and str(c) != str(cols[name_col_guess_idx])),
                     min(3, len(cols) - 1)),
            )

            name_col = cols[name_col_guess_idx]
            addr_col = cols[addr_col_guess_idx]

            lat_col_guess = next((c for c in cols if "위도" in str(c) or str(c).lower() == "lat"), None)
            lng_col_guess = next((c for c in cols if "경도" in str(c) or str(c).lower() in ("lng", "lon")), None)
            has_coords = bool(lat_col_guess and lng_col_guess)

            # 대상명·주소 열과 좌표 처리방식은 앱이 자동 결정한다.
            coord_mode = "file" if has_coords else "geocode"
            verify_tol_km = 1.0


            n_targets = len(df)
            if purpose == "inspect":
                available_team_days = len(inspect_dates) * int(inspect_teams)
                daily_target = math.ceil(n_targets / available_team_days) if available_team_days else 0
                inspection_only_capacity = math.floor(float(inspect_daily_hours) * 60 / int(inspect_minutes))

                st.markdown("**✅ 설정 결과 · 예방검사**")
                pc1, pc2, pc3, pc4 = st.columns(4)
                pc1.metric("전체 대상", f"{n_targets}개소")
                pc2.metric("검사 가능일", f"{len(inspect_dates)}일")
                pc3.metric("전체 가용량", f"{available_team_days}팀 일")
                pc4.metric("팀당 하루 권장량", f"{daily_target}개소" if daily_target else "계산 불가")

                if not available_team_days:
                    st.error("검사기간과 검사 가능 요일을 확인해주세요. 현재 배정 가능한 날짜가 없습니다.")
                elif daily_target > inspection_only_capacity:
                    st.warning(
                        f"현재 조건에서는 팀당 하루 최소 {daily_target}개소가 필요하지만, "
                        f"검사시간만 계산한 이론상 최대량은 {inspection_only_capacity}개소입니다. "
                        "이동시간까지 고려하면 기한 내 완료가 어려울 수 있으므로 팀 수·검사일·하루 가능시간을 늘려주세요."
                    )
                else:
                    st.success(
                        f"팀당 하루 평균 {n_targets / available_team_days:.1f}개소, "
                        f"권장 {daily_target}개소씩 배정하면 기한 내 검사가 가능합니다. "
                        "실제 노선 생성 시 이동시간을 함께 확인하세요."
                    )
                st.caption(
                    "이 계산에는 개인정보가 필요하지 않습니다. 대상 파일은 대상명과 주소만 사용하고 "
                    "담당자 이름·전화번호·검사결과는 업로드하지 마세요."
                )
            elif purpose == "hydrant":
                base_count, extra_count = divmod(n_targets, max(int(hydrant_member_count), 1))
                per_person_text = (f"{base_count}~{base_count + 1}개" if extra_count else f"{base_count}개")
                average_per_duty = n_targets / max(int(hydrant_member_count) * int(hydrant_workdays), 1)
                st.markdown("**✅ 설정 결과 · 월간 지리조사**")
                hp1, hp2, hp3, hp4 = st.columns(4)
                hp1.metric("전체 소화전", f"{n_targets}개")
                hp2.metric("조사 인원", f"{int(hydrant_member_count)}명")
                hp3.metric("개인별 월 담당", per_person_text)
                hp4.metric("당번 1회 평균", f"{average_per_duty:.1f}개/인")
                st.caption(
                    f"차량 {int(hydrant_vehicle_count)}대 · 월 당번 {int(hydrant_workdays)}일 · "
                    f"기본 노선 {int(hydrant_target_min)}분 · 최대 {int(hydrant_max_min)}분. "
                    "좌표 확정 후 같은 차량 팀원의 담당 구역이 인접하도록 자동 배정합니다."
                )
            elif purpose == "season":
                st.markdown("**✅ 설정 결과**")
                sp1, sp2, sp3, sp4 = st.columns(4)
                sp1.metric("관할 전체 대상", f"{n_targets}개소")
                sp2.metric("순찰 기간", f"{period_days}일")
                sp3.metric("1회 최대", f"{int(season_target_min)}분")
                sp4.metric("편도 제한", f"{season_oneway_limit:g}{'분' if season_limit_basis == '편도시간(분)' else 'km'}")
                st.caption(
                    f"{season_actor}/{season_vehicle} · 전 대상 균등 순환. "
                    f"실제 도로의 {season_limit_basis} 기준을 넘는 대상은 소방공무원 차량 코스에서 분리하여 "
                    f"{season_delegate} 대상으로 안내합니다."
                )
            elif purpose == "other":
                st.markdown("**✅ 설정 결과 · 지휘관 현장방문**")
                st.caption(
                    f"{visit_purpose} · {n_targets}개소 · {commander_route_mode}. "
                    "목표시간이나 원거리 제외 없이 실제 도로거리순으로 전 대상을 연결합니다."
                )

            saved_coords = st.session_state.get("coords_df")
            coords_complete = (saved_coords is not None and len(saved_coords) == n_targets
                               and saved_coords["위도"].notna().all()
                               and saved_coords["경도"].notna().all())
            find_coords = False
            if st.session_state.get("coord_future") is not None:
                st.info("🔍 좌표 검색 중입니다. 위의 ③ 대상 좌표 우선 확인에서 완료 상태를 확인하세요.")
            elif saved_coords is None:
                st.warning("⚠️ 위의 ③ 대상 좌표 우선 확인에서 좌표를 먼저 검색해주세요.")

        # ---- 1단계: 좌표 확정 ---------------------------------------------------
        if find_coords:
            rows = []
            prog = st.progress(0.0, text="주소로 좌표를 찾는 중...")
            n = len(df)
            for i, (_, row) in enumerate(df.iterrows()):
                nm, ad = str(row[name_col]), str(row[addr_col])

                file_lat = file_lng = None
                if lat_col_guess and lng_col_guess:
                    try:
                        file_lat, file_lng = float(row[lat_col_guess]), float(row[lng_col_guess])
                        if math.isnan(file_lat) or math.isnan(file_lng):
                            file_lat = file_lng = None
                    except (TypeError, ValueError):
                        file_lat = file_lng = None

                if coord_mode == "file" and file_lat is not None:
                    rows.append({"대상명": nm, "주소": ad, "위도": file_lat, "경도": file_lng,
                                 "상태": "파일 좌표", "비고": "파일에 있던 좌표를 그대로 사용"})
                else:
                    lat, lng, used_q, used_why, tried = geocode_with_fallback(ad, nm)
                    if lat is None and file_lat is not None:
                        rows.append({"대상명": nm, "주소": ad, "위도": file_lat, "경도": file_lng,
                                     "상태": "⚠ 파일 좌표로 대체",
                                     "비고": "주소로는 못 찾아 파일 좌표를 사용했습니다"})
                    elif lat is None:
                        rows.append({"대상명": nm, "주소": ad, "위도": None, "경도": None,
                                     "상태": "❌ 실패",
                                     "비고": "시도: " + " / ".join(tried)})
                    else:
                        gap_note = ""
                        if coord_mode == "verify" and file_lat is not None:
                            gap = haversine_km((file_lat, file_lng), (lat, lng))
                            if gap > verify_tol_km:
                                gap_note = f" (파일 좌표와 {gap:.1f}km 차이 — API 좌표로 교체)"
                            else:
                                lat, lng = file_lat, file_lng
                        rows.append({
                            "대상명": nm, "주소": ad, "위도": lat, "경도": lng,
                            "상태": "✅ 확인" if used_why == "원본 주소" else "🔧 주소 보정 후 확인",
                            "비고": ("" if used_why == "원본 주소" else f"{used_why} → {used_q}") + gap_note,
                        })
                prog.progress((i + 1) / n, text=f"주소로 좌표를 찾는 중... ({i+1}/{n})")
            prog.empty()
            st.session_state["coords_df"] = pd.DataFrame(rows)
            st.session_state.pop("route_results", None)   # 좌표가 바뀌었으니 이전 노선 결과는 폐기
            if st.session_state["coords_df"]["위도"].notna().all() and st.session_state["coords_df"]["경도"].notna().all():
                st.rerun()

        coords_df = st.session_state.get("coords_df")

        if coords_df is not None:
            st.write("")
            coord_detail_col, _ = st.columns([1, 1])
            with coord_detail_col:
                coord_detail_box = st.expander("과정 보기 · 좌표 결과 확인 및 수정", expanded=False)
            with coord_detail_box:
                ok_n = int(coords_df["위도"].notna().sum())
                fail_n = int(coords_df["위도"].isna().sum())
                k1, k2, k3 = st.columns(3)
                k1.metric("전체", f"{len(coords_df)}")
                k2.metric("좌표 확보", f"{ok_n}")
                k3.metric("좌표 없음", f"{fail_n}")

                if fail_n:
                    st.error(f"❌ {fail_n}건은 좌표를 찾지 못했습니다. 아래 표의 **위도·경도 칸에 직접 입력**하시면 "
                             "노선 생성에 포함됩니다. (네이버·카카오 지도에서 해당 지점을 찍고 좌표를 확인해 넣으시면 됩니다.)")
                else:
                    st.success("✅ 모든 대상의 좌표가 확보되었습니다. 아래 2단계로 진행하세요.")

                st.caption("위도·경도 칸은 직접 고칠 수 있습니다. 수정하면 그 값이 노선 생성에 그대로 쓰입니다.")
                edited = st.data_editor(
                    coords_df, use_container_width=True, hide_index=True, num_rows="fixed",
                    key="coords_editor",
                    column_config={
                        "대상명": st.column_config.TextColumn(disabled=True, width="medium"),
                        "주소": st.column_config.TextColumn(disabled=True, width="large"),
                        "위도": st.column_config.NumberColumn(format="%.6f", help="예: 35.919000"),
                        "경도": st.column_config.NumberColumn(format="%.6f", help="예: 128.283000"),
                        "상태": st.column_config.TextColumn(disabled=True, width="small"),
                        "비고": st.column_config.TextColumn(disabled=True, width="large"),
                    },
                )
                st.session_state["coords_df"] = edited

                st.download_button(
                    "📥 확정된 좌표 CSV로 저장 (다음엔 이 파일을 올리면 좌표 찾기 없이 바로 진행)",
                    data=edited.to_csv(index=False).encode("utf-8-sig"),
                    file_name="확정좌표.csv", mime="text/csv",
                )

            st.write("")
            with st.container(border=True):
                st.markdown("**노선 생성**")
                ready = edited["위도"].notna() & edited["경도"].notna()
                n_ready = int(ready.sum())

                if candidate_k:
                    est_calls = n_ready * candidate_k + n_ready
                else:
                    est_calls = n_ready * (n_ready + 1) // 2 + n_ready
                est_sec = int(est_calls * 0.25)
                if n_ready < len(edited):
                    st.warning(f"좌표가 없는 {len(edited) - n_ready}건은 노선에서 제외됩니다.")

                route_complete = bool(st.session_state.get("route_results"))
                route_action_col, _, _ = st.columns(3)
                with route_action_col:
                    if route_complete:
                        st.markdown(
                            '<div class="paseru-complete-action">✅ 노선 생성 완료 · 아래에서 결과 확인</div>',
                            unsafe_allow_html=True,
                        )
                        with st.expander("과정 보기 · 노선 다시 계산", expanded=False):
                            st.write(f"좌표 {n_ready}개 · 예상 API 호출 약 {est_calls:,}회 · 최대 {max_calls:,}회")
                            run = st.button("노선 다시 생성", type="secondary", use_container_width=True)
                    else:
                        with st.expander("과정 보기 · 예상 계산량", expanded=False):
                            st.write(f"좌표 {n_ready}개 · 예상 API 호출 약 {est_calls:,}회 · 최대 {max_calls:,}회")
                            st.caption(f"예상 계산시간 약 {est_sec // 60}분 {est_sec % 60}초")
                        run = st.button("노선 생성 시작", type="primary",
                                        disabled=(not has_keys() or n_ready == 0), use_container_width=True)
        else:
            run = False
            st.info("먼저 위의 **③ 대상 좌표 우선 확인**에서 좌표를 확정해 주세요.")

        if run:
            # ---- 중단 장치 ----------------------------------------------------
            # ① 수동 중단: 아래 '중단' 버튼을 누르면 Streamlit이 새로 실행되면서
            #    지금 돌고 있는 계산이 즉시 멈춘다.
            # ② 자동 중단: 호출 수가 한도(max_calls)를 넘으면 그때까지 만든 노선만 남기고 멈춘다.
            stop_box = st.container()
            with stop_box:
                st.button("⛔ 계산 중단", key="stop_btn", type="secondary",
                          help="지금 진행 중인 노선 생성을 즉시 멈춥니다.")

            call_counter = {"n": 0}
            limit_hit = {"v": False}

            def over_limit():
                if call_counter["n"] >= max_calls:
                    limit_hit["v"] = True
                    return True
                return False

            # 1) 소방서 좌표
            with st.spinner("소방서 좌표 확인 중..."):
                s_lat, s_lng = geocode_address(station_address)
            if s_lat is None:
                st.error("소방서 주소 지오코딩에 실패했습니다. 주소를 확인해주세요.")
                st.stop()
            station = {"name": station_name, "lat": s_lat, "lng": s_lng}

            # 2) 1단계에서 확정한 좌표를 그대로 사용 (여기서는 지오코딩을 하지 않는다)
            points = []
            for _, r in edited.iterrows():
                try:
                    lat, lng = float(r["위도"]), float(r["경도"])
                except (TypeError, ValueError):
                    continue
                if math.isnan(lat) or math.isnan(lng):
                    continue
                points.append({"name": str(r["대상명"]), "address": str(r["주소"]),
                               "lat": lat, "lng": lng})

            if not points:
                st.error("좌표가 있는 대상이 없습니다. 1단계에서 좌표를 확정해 주세요.")
                st.stop()

            def bump(total_hint=len(points)):
                call_counter["n"] += 1
            # 3) 용도별 원거리 판정
            if purpose == "hydrant":
                normal_points = allocate_hydrants_to_members(points, station, hydrant_members)
                far_points = []
            elif purpose == "other":
                # 지휘관 현장방문은 거리가 멀어도 전 대상을 반드시 포함한다.
                normal_points = points
                far_points = []
            else:
                long_progress = st.progress(0.0, text="소방서 기준 실도로거리 확인 중...")

                def bump(total_hint=len(points)):
                    call_counter["n"] += 1
                    long_progress.progress(min(call_counter["n"] / max(total_hint, 1), 1.0),
                                           text=f"실제 도로거리 API 호출 중... "
                                                f"({call_counter['n']:,}/{max_calls:,}회)")

                if purpose == "season":
                    if season_limit_basis == "편도시간(분)":
                        normal_points, far_points = separate_long_time(
                            points, station, float(season_oneway_limit), season_delegate,
                            on_call=bump, should_stop=over_limit,
                        )
                    else:
                        normal_points, far_points = separate_long_distance(
                            points, station, float(season_oneway_limit), on_call=bump,
                            save_calls=False, should_stop=over_limit,
                        )
                        far_points = [{**point, "권장수행": season_delegate} for point in far_points]
                else:
                    normal_points, far_points = separate_long_distance(
                        points, station, long_threshold, on_call=bump, save_calls=bool(candidate_k),
                        should_stop=over_limit,
                    )
                long_progress.empty()

            # 4) 노선 편성
            build_progress = st.empty()

            def bump_build():
                call_counter["n"] += 1
                build_progress.text(f"실도로 기준 노선 편성 중... "
                                    f"(API 호출 {call_counter['n']:,}/{max_calls:,}회)")

            if purpose == "hydrant":
                routes = []
                unassigned = []
                for vehicle_no in range(1, int(hydrant_vehicle_count) + 1):
                    vehicle_points = [p for p in normal_points if p.get("vehicle_no") == vehicle_no]
                    if not vehicle_points:
                        continue
                    vehicle_routes, vehicle_unassigned = build_routes(
                        vehicle_points, station, "target_time", 100,
                        None, None, int(hydrant_target_min),
                        None, basis="time", on_call=bump_build,
                        candidate_k=candidate_k, should_stop=over_limit,
                        service_min_per_stop=int(hydrant_inspection_min),
                    )
                    routes.extend(vehicle_routes)
                    unassigned.extend(vehicle_unassigned)
            elif purpose == "other":
                commander_route_count = (int(commander_vehicle_count)
                                         if commander_route_mode == "여러 차량으로 균등 분할" else 1)
                commander_per_route = max(1, math.ceil(len(normal_points) / commander_route_count))
                routes, unassigned = build_routes(
                    normal_points, station, "fixed", commander_per_route,
                    None, None, None, commander_route_count,
                    basis="distance", on_call=bump_build,
                    candidate_k=candidate_k, should_stop=over_limit,
                )
            else:
                routes, unassigned = build_routes(
                    normal_points, station, mode, max_per_route,
                    seg_max_km, seg_max_min, target_min_high,
                    max_routes_cap or None, basis=basis, on_call=bump_build,
                    candidate_k=candidate_k, should_stop=over_limit,
                    service_min_per_stop=(int(season_stop_min) if purpose == "season" else 0),
                )
            build_progress.empty()

            if limit_hit["v"]:
                st.warning(
                    f"⛔ API 호출 한도({max_calls:,}회)에 도달해 노선 편성을 중단했습니다. "
                    f"그때까지 편성된 {len(routes)}개 노선은 아래에 그대로 표시됩니다. "
                    "한도를 늘리거나 'API 호출 절약'을 켜고 다시 실행해 보세요."
                )

            if unassigned:
                for p in unassigned:
                    km, _ = real_leg(station, p)
                    far_points.append({**p, "도로거리_km": round(km, 1)})

            # 용도별 부가 정보
            team_info = ""
            if purpose == "hydrant":
                base_count, extra_count = divmod(len(points), max(int(hydrant_member_count), 1))
                count_text = (f"{base_count}~{base_count + 1}개" if extra_count else f"{base_count}개")
                team_info = (f" · {hydrant_member_count}명 개인별 {count_text}"
                             f" · 차량 {hydrant_vehicle_count}대 · 월 {hydrant_workdays}근무일")
            elif purpose == "season":
                limit_unit = "분" if season_limit_basis == "편도시간(분)" else "km"
                team_info = (f" · 관할 전체 대상 균등 순환 · {season_actor}/{season_vehicle}"
                             f" · 1회 최대 {int(season_target_min)}분"
                             f" · 편도 {season_oneway_limit:g}{limit_unit} 제한")
            elif purpose == "other":
                team_info = (f" · {visit_purpose} · {commander_route_mode}"
                             + (f"({int(commander_vehicle_count)}대)"
                                if commander_route_mode == "여러 차량으로 균등 분할" else ""))
            elif purpose == "inspect":
                available_team_days = len(inspect_dates) * int(inspect_teams)
                assigned_targets = sum(len(route) for route in routes)
                daily_target = math.ceil(assigned_targets / available_team_days) if available_team_days else 0
                team_info = (f" · 검사 가능일 {len(inspect_dates)}일 · {inspect_teams}팀"
                             + (f" · 팀당 하루 최소 {daily_target}개소" if daily_target else ""))
            elif purpose == "guard" and guard_repeat_label == "매일 같은 코스 반복" and guard_rounds:
                total_runs = int(guard_rounds.replace("회", "")) * period_days
                team_info = f" · 매일 같은 코스로 하루 {guard_rounds} 반복({period_days}일간 총 {total_runs}회)"
            elif purpose == "guard":
                team_info = f" · 매일 다른 코스로 순환({period_days}일간 {len(routes)}개 노선 배정)"

            far_word = "편도 기준 초과" if purpose == "season" else "장거리 별도"
            st.success(f"[{purpose_label}] 총 {len(routes)}개 노선, {sum(len(r) for r in routes)}개소 배정 완료 "
                       f"({far_word} {len(far_points)}개소){team_info}")
            # 5) 확정 노선의 구간별 실도로거리·경로좌표
            route_results = []
            total_calls = sum(len(r) + 1 for r in routes)
            call_progress = st.progress(0.0, text="노선별 실도로 경로 확정 중...")
            done = 0
            for ri, route in enumerate(routes):
                legs = []
                cur = station
                acc_km = 0.0
                acc_min = 0.0
                all_path = []
                for p in route:
                    km, mins, path = road_route(cur["lat"], cur["lng"], p["lat"], p["lng"])
                    if km is None:
                        km = haversine_km((cur["lat"], cur["lng"]), (p["lat"], p["lng"])) * ROAD_FACTOR
                        mins = km / AVG_SPEED_KMH * 60
                        path = [(cur["lat"], cur["lng"]), (p["lat"], p["lng"])]
                    service_min = (int(hydrant_inspection_min) if purpose == "hydrant" else
                                   int(season_stop_min) if purpose == "season" else 0)
                    legs.append({"from": cur["name"], "to": p["name"], "to_address": p.get("address", ""),
                                 "km": km, "min": mins, "inspection_min": service_min,
                                 "assigned_to": p.get("assigned_to", ""),
                                 "vehicle_no": p.get("vehicle_no"),
                                 "lat": p["lat"], "lng": p["lng"]})
                    acc_km += km
                    acc_min += mins + service_min
                    all_path += path
                    cur = p
                    done += 1
                    call_progress.progress(min(done / max(total_calls, 1), 1.0), text="노선별 실도로 경로 확정 중...")
                back_km, back_min, back_path = road_route(cur["lat"], cur["lng"], station["lat"], station["lng"])
                if back_km is None:
                    back_km = haversine_km((cur["lat"], cur["lng"]), (station["lat"], station["lng"])) * ROAD_FACTOR
                    back_min = back_km / AVG_SPEED_KMH * 60
                    back_path = [(cur["lat"], cur["lng"]), (station["lat"], station["lng"])]
                acc_km += back_km
                acc_min += back_min
                all_path += back_path
                done += 1
                call_progress.progress(min(done / max(total_calls, 1), 1.0), text="노선별 실도로 경로 확정 중...")

                route_results.append({
                    "route_no": ri + 1, "stops": route, "legs": legs,
                    "vehicle_no": route[0].get("vehicle_no") if route else None,
                    "assigned_members": sorted({p.get("assigned_to", "") for p in route if p.get("assigned_to")}),
                    "back_km": back_km, "back_min": back_min,
                    "total_km": acc_km, "total_min": acc_min, "path": all_path,
                })
            call_progress.empty()

            if purpose == "hydrant":
                route_counts = {
                    vehicle_no: sum(1 for result in route_results if result.get("vehicle_no") == vehicle_no)
                    for vehicle_no in range(1, int(hydrant_vehicle_count) + 1)
                }
                over_vehicles = {v: count for v, count in route_counts.items() if count > int(hydrant_workdays)}
                if over_vehicles:
                    detail = ", ".join(f"{v}호차 {count}개 노선" for v, count in over_vehicles.items())
                    suggested = min(
                        int(hydrant_max_min),
                        math.ceil(int(hydrant_target_min) * max(over_vehicles.values()) / int(hydrant_workdays) / 10) * 10,
                    )
                    st.warning(
                        f"월 {int(hydrant_workdays)}번의 당번일을 초과하는 차량이 있습니다: {detail}. "
                        f"우선 목표시간을 약 {suggested}분으로 늘려 다시 편성해 보세요. "
                        f"{int(hydrant_max_min)}분에서도 10개 이내가 되지 않으면 인원 또는 차량 편성을 조정해야 합니다."
                    )
                else:
                    detail = " · ".join(f"{v}호차 {count}개 노선" for v, count in route_counts.items())
                    st.success(f"월 {int(hydrant_workdays)}번의 당번일 안에 전수조사가 가능합니다. {detail}")
            elif purpose == "season":
                required_routes = len(route_results)
                if required_routes > period_days:
                    st.warning(
                        f"전 대상을 누락 없이 순찰하려면 1회 최대 {int(season_target_min)}분 기준으로 "
                        f"최소 {required_routes}일이 필요하지만 "
                        f"설정한 기간은 {period_days}일입니다. 기간을 {required_routes}일 이상으로 늘리거나 "
                        "1회 최대시간을 늘려 다시 편성해주세요."
                    )
                elif required_routes:
                    base_runs, extra_runs = divmod(period_days, required_routes)
                    for route_index, result in enumerate(route_results):
                        result["period_runs"] = base_runs + (1 if route_index < extra_runs else 0)
                    min_runs = base_runs
                    max_runs = base_runs + (1 if extra_runs else 0)
                    run_text = f"{min_runs}회" if min_runs == max_runs else f"{min_runs}~{max_runs}회"
                    st.success(
                        f"일반 순찰 대상 {sum(len(result['stops']) for result in route_results)}개소를 "
                        f"{required_routes}개 코스로 나누었습니다. {period_days}일 동안 코스를 차례로 순환하면 "
                        f"각 코스와 소속 대상은 예상 {run_text} 방문합니다."
                    )
                    if far_points:
                        st.info(
                            f"편도 기준을 넘는 {len(far_points)}개소는 출동 공백을 줄이기 위해 "
                            f"{season_delegate} 대상으로 별도 안내합니다."
                        )

            st.session_state["station"] = station
            st.session_state["route_results"] = route_results
            st.session_state["far_points"] = far_points
            st.session_state["meta"] = {
                "title": patrol_title, "purpose": purpose_label, "vehicle": vehicle,
                "period": (f"{period_start:%Y-%m-%d} ~ {period_end:%Y-%m-%d} "
                           f"(검사 가능일 {len(inspect_dates)}일)" if purpose == "inspect" else
                           f"{period_start:%Y년 %m월} · 월 당번 {int(hydrant_workdays)}일" if purpose == "hydrant" else
                           f"{period_start:%Y-%m-%d}" if purpose == "other" else
                           f"{start_dt:%Y-%m-%d %H:%M} ~ {end_dt:%Y-%m-%d %H:%M} ({period_days}일간)"),
                "period_label": ("검사기간" if purpose == "inspect" else
                                 "조사월" if purpose == "hydrant" else
                                 "방문일" if purpose == "other" else "순찰기간"),
                "basis": basis_label, "route_prefix": route_prefix, "team_info": team_info.strip(" ·"),
                "target_min": target_min,
            }

    # ----------------------------------------------------------------------------

with page_results:
    # 결과 표시
    # ----------------------------------------------------------------------------
    if st.session_state.get("stop_btn"):
        st.info("⛔ 계산을 중단했습니다. (중단 시점까지의 계산 결과는 저장되지 않습니다. "
                "직전에 완료된 결과가 있으면 아래에 그대로 남아 있습니다.)")

    if "route_results" in st.session_state:
        station = st.session_state["station"]
        route_results = st.session_state["route_results"]
        far_points = st.session_state["far_points"]
        meta = st.session_state.get("meta", {})

        st.write("")
        st.header("📍 노선 생성 결과")
        if meta:
            st.caption(f"**{meta.get('title','')}** · {meta.get('purpose','')} · 기준: {meta.get('basis','')} · "
                       f"{meta.get('period_label', '순찰기간')} {meta.get('period','')} · 차량: {meta.get('vehicle','')}"
                       + (f" · {meta['team_info']}" if meta.get("team_info") else ""))

        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("생성 노선 수", f"{len(route_results)}")
        m2.metric("전체 방문지", f"{sum(len(r['stops']) for r in route_results)}")
        m3.metric("총 이동거리(km)", f"{sum(r['total_km'] for r in route_results):.1f}")
        m4.metric("노선 평균시간", f"{sum(r['total_min'] for r in route_results) / max(len(route_results), 1):.0f}분")
        m5.metric("편도 기준 초과" if meta.get("purpose") == "③ 계절순찰" else "원거리 분리 대상",
                  f"{len(far_points)}")
        if meta.get("purpose") == "③ 계절순찰":
            visit_runs = [r.get("period_runs", 0) for r in route_results if r.get("period_runs")]
            if visit_runs:
                repeat_text = (f"{min(visit_runs)}회" if min(visit_runs) == max(visit_runs)
                               else f"{min(visit_runs)}~{max(visit_runs)}회")
                st.info(f"🔁 설정 기간 동안 같은 대상의 예상 방문 횟수: {repeat_text}")

        # ---- 담당 조 · 조원 입력(화면에서 직접 입력 → 엑셀에 그대로 반영) ----
        with st.expander("선택 사항 · 노선별 담당 조와 조원 입력", expanded=False):
            st.caption("필요한 경우에만 입력하세요. 입력 내용은 최종 엑셀 파일에 반영됩니다.")
            for rr in route_results:
                tc1, tc2, tc3 = st.columns([0.8, 1.2, 2])
                with tc1:
                    st.markdown(f"**노선 {rr['route_no']}**")
                with tc2:
                    st.text_input("담당 조 이름", key=f"team_name_{rr['route_no']}",
                                  placeholder="예) 가천1팀1조", label_visibility="collapsed")
                with tc3:
                    st.text_input("조원", key=f"team_members_{rr['route_no']}",
                                  placeholder="조원 예) 홍길동, 이순신", label_visibility="collapsed")

        # ---- 엑셀(xlsx) 다운로드: 노선 1개 = 1행, 경유지가 옆으로 펼쳐지는 가로형 ----
        def build_wide_excel(station, route_results, far_points, meta):
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "순찰노선"

            prefix = (meta.get("route_prefix") or "").strip() or station["name"]
            base_cols = ["연번", "부서명", "노선이름", "구분", "담당 조", "조원", "출발지"]
            max_stops = max((len(rr["legs"]) for rr in route_results), default=0)

            thin = Side(style="thin", color="9AA59D")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            head_fill = PatternFill("solid", fgColor="F4DDD8")
            head_font = Font(bold=True, size=10)
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # 제목 줄
            total_cols = len(base_cols) + max_stops * 3 + 3
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(total_cols, 1))
            title_cell = ws.cell(row=1, column=1, value=f"{meta.get('title', '순찰노선')}   "
                                                        f"[{meta.get('purpose', '')} · {meta.get('vehicle', '')} · "
                                                        f"{meta.get('period', '')}]")
            title_cell.font = Font(bold=True, size=13)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            HEAD1, HEAD2, DATA0 = 2, 3, 4

            for i, name in enumerate(base_cols, start=1):
                ws.merge_cells(start_row=HEAD1, start_column=i, end_row=HEAD2, end_column=i)
                ws.cell(row=HEAD1, column=i, value=name)

            col = len(base_cols) + 1
            for i in range(1, max_stops + 1):
                ws.merge_cells(start_row=HEAD1, start_column=col, end_row=HEAD1, end_column=col + 2)
                ws.cell(row=HEAD1, column=col, value=f"경유{i}")
                ws.cell(row=HEAD2, column=col, value="대상명(주소)")
                ws.cell(row=HEAD2, column=col + 1, value="거리(km)")
                ws.cell(row=HEAD2, column=col + 2, value="누적(km)")
                col += 3

            for label in ("귀소", "노선거리(km)", "순찰 총 소요시간(분)"):
                ws.merge_cells(start_row=HEAD1, start_column=col, end_row=HEAD2, end_column=col)
                ws.cell(row=HEAD1, column=col, value=label)
                col += 1
            last_col = col - 1

            for r_ in (HEAD1, HEAD2):
                for c_ in range(1, last_col + 1):
                    cell = ws.cell(row=r_, column=c_)
                    cell.fill = head_fill
                    cell.font = head_font
                    cell.alignment = center
                    cell.border = border

            r = DATA0
            for rr in route_results:
                no = rr["route_no"]
                c = 1
                ws.cell(row=r, column=c, value=no); c += 1
                ws.cell(row=r, column=c, value=station["name"]); c += 1
                ws.cell(row=r, column=c, value=f"{prefix}노선{no}"); c += 1
                ws.cell(row=r, column=c, value="근거리"); c += 1
                auto_team = f"{rr.get('vehicle_no')}호차" if rr.get("vehicle_no") else ""
                auto_members = ", ".join(rr.get("assigned_members") or [])
                ws.cell(row=r, column=c, value=(st.session_state.get(f"team_name_{no}", "") or auto_team)); c += 1
                ws.cell(row=r, column=c, value=(st.session_state.get(f"team_members_{no}", "") or auto_members)); c += 1
                ws.cell(row=r, column=c, value=station["name"]); c += 1
                acc_km = 0.0
                for leg in rr["legs"]:
                    acc_km += leg["km"]
                    ws.cell(row=r, column=c, value=stop_label(leg["to"], leg.get("to_address"))); c += 1
                    ws.cell(row=r, column=c, value=round(leg["km"], 1)); c += 1
                    ws.cell(row=r, column=c, value=round(acc_km, 1)); c += 1
                c += (max_stops - len(rr["legs"])) * 3  # 경유지 수가 적은 노선은 빈칸 패딩
                ws.cell(row=r, column=c, value=station["name"]); c += 1
                ws.cell(row=r, column=c, value=round(rr["total_km"], 1)); c += 1
                ws.cell(row=r, column=c, value=round(rr["total_min"])); c += 1
                r += 1

            for idx, p in enumerate(far_points, start=1):
                c = 1
                ws.cell(row=r, column=c, value=f"장거리{idx}"); c += 1
                ws.cell(row=r, column=c, value=station["name"]); c += 1
                ws.cell(row=r, column=c, value="-"); c += 1
                ws.cell(row=r, column=c, value="원거리"); c += 1
                c += 2  # 담당 조 · 조원 빈칸
                ws.cell(row=r, column=c, value=station["name"]); c += 1
                ws.cell(row=r, column=c, value=stop_label(p["name"], p.get("address"))); c += 1
                ws.cell(row=r, column=c, value=p.get("도로거리_km", "")); c += 1
                r += 1

            for row_cells in ws.iter_rows(min_row=DATA0, max_row=r - 1, min_col=1, max_col=last_col):
                for cell in row_cells:
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

            for i in range(1, last_col + 1):
                width = 24 if (i > len(base_cols) and (i - len(base_cols) - 1) % 3 == 0) else 13
                ws.column_dimensions[get_column_letter(i)].width = width
            ws.freeze_panes = ws.cell(row=DATA0, column=len(base_cols) + 1)

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        safe_title = re.sub(r'[\\/:*?"<>|]', "_", meta.get("title", "순찰노선")) or "순찰노선"
        with st.container(border=True):
            card_title(5, "최종 출력 · 현장 전달")
            st.success("✅ 노선 생성이 완료되었습니다. 필요한 형식으로 최종 결과를 내려받으세요.")
            st.caption("엑셀 순찰표를 먼저 저장한 뒤, 현장 전달 방식에 따라 링크·QR·인쇄 자료를 선택하세요.")

            out_left, out_right = st.columns(2)
            with out_left:
                st.download_button(
                    "📥 최종 순찰표 엑셀 받기",
                    data=build_wide_excel(station, route_results, far_points, meta),
                    file_name=f"{safe_title}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with out_right:
                st.download_button(
                    "🔗 카카오맵 경로·링크 엑셀 받기",
                    data=build_route_links_excel(station, route_results),
                    file_name=f"{safe_title}_노선별_경로와_링크.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            qr_left, qr_right = st.columns(2)
            with qr_left:
                st.download_button(
                    "📦 전체 노선 QR 묶음 받기",
                    data=build_qr_zip(station, route_results),
                    file_name=f"{safe_title}_QR코드_전체.zip",
                    mime="application/zip",
                    use_container_width=True,
                )
            with qr_right:
                st.download_button(
                    "🖨 전체 노선 QR 인쇄문서 받기",
                    data=build_printable_qr_html(station, route_results, meta),
                    file_name=f"{safe_title}_QR인쇄.html",
                    mime="text/html",
                    use_container_width=True,
                )
            st.caption("인쇄용 문서를 열고 오른쪽 위의 ‘인쇄하기’를 누르면 노선별로 A4 한 장씩 출력됩니다.")

        target_min_ref = meta.get("target_min")
        if target_min_ref:
            over = [r for r in route_results if r["total_min"] > target_min_ref]
            if over:
                st.warning(
                    f"⏱ 목표 {target_min_ref}분을 넘는 노선이 {len(over)}개 있습니다 "
                    f"(노선 {', '.join(str(r['route_no']) for r in over)}). "
                    "노선당 구간 수를 줄이거나 목표시간을 늘려 다시 편성해 보세요."
                )
            else:
                st.success(f"⏱ 모든 노선이 목표 {target_min_ref}분 이내입니다.")

        for rr in route_results:
            team_name = st.session_state.get(f"team_name_{rr['route_no']}", "")
            over_mark = " ⏱초과" if target_min_ref and rr["total_min"] > target_min_ref else ""
            hydrant_label = ""
            if meta.get("purpose") == "⑤ 지리조사(센터용)" and rr.get("vehicle_no"):
                members = ", ".join(rr.get("assigned_members") or [])
                hydrant_label = f" · {rr['vehicle_no']}호차" + (f" · {members}" if members else "")
            season_runs = (f" · 기간 중 {rr.get('period_runs', 0)}회 예상"
                           if meta.get("purpose") == "계절순찰" and rr.get("period_runs") else "")
            head = (f"노선 {rr['route_no']}" + hydrant_label + (f" · {team_name}" if team_name else "") +
                    f" — {len(rr['stops'])}개소 · 총 {rr['total_km']:.1f}km · 약 {rr['total_min']:.0f}분"
                    + season_runs + over_mark)
            with st.expander(f"🔎 상세 과정 보기 · {head}", expanded=False):
                col1, col2 = st.columns([1, 1])

                with col1:
                    rows = []
                    for i, leg in enumerate(rr["legs"], start=1):
                        row = {
                            "순번": str(i), "지점": stop_label(leg["to"], leg.get("to_address")),
                            "구간거리(km)": round(leg["km"], 1), "구간시간(분)": round(leg["min"]),
                        }
                        if meta.get("purpose") == "⑤ 지리조사(센터용)":
                            row["담당"] = leg.get("assigned_to", "")
                            row["조사시간(분)"] = leg.get("inspection_min", 0)
                        rows.append(row)
                    rows.append({"순번": "", "지점": f"복귀 ({station['name']})",
                                 "구간거리(km)": round(rr["back_km"], 1),
                                 "구간시간(분)": round(rr["back_min"])})
                    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

                    st.markdown("**🟨 카카오맵 — 전체 순찰코스와 QR코드**")
                    kakao_links = kakao_route_links(station, rr["legs"])
                    st.download_button(
                        f"🖨 노선 {rr['route_no']} QR 인쇄용 문서",
                        data=build_printable_qr_html(station, [rr], meta),
                        file_name=f"{safe_title}_노선_{rr['route_no']}_QR인쇄.html",
                        mime="text/html",
                        key=f"route_print_{rr['route_no']}",
                        use_container_width=True,
                    )
                    for li, (kurl, origin, destinations) in enumerate(kakao_links, start=1):
                        suffix = "" if len(kakao_links) == 1 else f" ({li}/{len(kakao_links)}구간)"
                        seq = " → ".join([origin["name"]] + [p["name"] for p in destinations])
                        qr_png = make_qr_png(kurl)

                        guide_col, qr_col = st.columns([1.35, 1])
                        with guide_col:
                            st.link_button(
                                f"🚗 카카오맵 전체 코스 길안내{suffix}",
                                kurl,
                                use_container_width=True,
                            )
                        with qr_col:
                            qr_box = (st.popover(f"📱 QR코드 보기{suffix}", use_container_width=True)
                                      if hasattr(st, "popover")
                                      else st.expander(f"📱 QR코드 보기{suffix}"))
                            with qr_box:
                                st.image(qr_png, caption="휴대폰 카메라로 스캔하세요.", width=260)
                                st.download_button(
                                    "QR코드 이미지 저장",
                                    data=qr_png,
                                    file_name=f"노선_{rr['route_no']}_카카오맵_QR_{li}.png",
                                    mime="image/png",
                                    key=f"kakao_qr_{rr['route_no']}_{li}",
                                    use_container_width=True,
                                )
                        st.caption(f"경로: {seq}")

                    if len(kakao_links) > 1:
                        st.caption(
                            f"※ 카카오맵은 경유지를 한 구간에 최대 {KAKAO_MAX_VIA}개까지 지원하므로 "
                            "긴 노선은 이어지는 구간으로 나눴습니다. 현장에서 순서대로 열어 주세요."
                        )
                    else:
                        st.caption(
                            "※ 소방서를 출발지와 최종 목적지로, 순찰 대상을 경유지 순서대로 "
                            "입력한 카카오맵 링크입니다. QR을 스캔하면 같은 코스가 열립니다."
                        )

                    with st.expander("지점별 개별 길안내(카카오맵) — 예비용", expanded=False):
                        for i, leg in enumerate(rr["legs"], start=1):
                            st.markdown(
                                f"- [{i}. {leg['to']} 길안내]({kakao_url(leg['to'], leg['lat'], leg['lng'])})"
                            )
                        st.markdown(
                            f"- [🚒 {station['name']} 귀소 길안내]"
                            f"({kakao_url(station['name'], station['lat'], station['lng'])})"
                        )
                        st.caption(
                            "전체 코스가 특정 휴대폰에서 열리지 않을 때 사용하세요. 각 버튼은 "
                            "현재 위치에서 선택한 다음 지점까지 안내합니다."
                        )

                with col2:
                    m = folium.Map(location=[station["lat"], station["lng"]], zoom_start=12)
                    if rr["path"]:
                        folium.PolyLine(rr["path"], color="#a33a3f", weight=4, opacity=0.85).add_to(m)

                    # 방문 순서를 지도 위에 숫자로 표시 (기본 핀 대신 번호 원)
                    for i, leg in enumerate(rr["legs"], start=1):
                        folium.Marker(
                            [leg["lat"], leg["lng"]],
                            tooltip=f"{i}. {leg['to']}",
                            icon=folium.DivIcon(
                                icon_size=(30, 30), icon_anchor=(15, 15),
                                html=(
                                    '<div style="background:#1f6fb2;color:#ffffff;'
                                    'width:26px;height:26px;border-radius:50%;'
                                    'border:2px solid #ffffff;box-shadow:0 1px 5px rgba(0,0,0,.45);'
                                    'display:flex;align-items:center;justify-content:center;'
                                    'font-family:sans-serif;font-weight:700;font-size:13px;'
                                    f'line-height:1;">{i}</div>'
                                ),
                            ),
                        ).add_to(m)

                    # 출발·복귀 지점은 눈에 띄게 빨간 '출발' 표식으로
                    folium.Marker(
                        [station["lat"], station["lng"]], tooltip=f"출발·복귀: {station['name']}",
                        icon=folium.DivIcon(
                            icon_size=(56, 26), icon_anchor=(28, 13),
                            html=(
                                '<div style="background:#a33a3f;color:#ffffff;'
                                'padding:3px 8px;border-radius:13px;border:2px solid #ffffff;'
                                'box-shadow:0 1px 5px rgba(0,0,0,.45);text-align:center;'
                                'font-family:sans-serif;font-weight:700;font-size:12px;'
                                'line-height:1.2;white-space:nowrap;">🚒 출발</div>'
                            ),
                        ),
                    ).add_to(m)
                    st_folium(m, height=350, use_container_width=True, key=f"map_{rr['route_no']}")

        if far_points:
            is_season_far = meta.get("purpose") == "③ 계절순찰"
            st.header("🚙 편도 기준 초과 대상" if is_season_far else "⚠️ 장거리 별도 대상")
            far_df = pd.DataFrame(far_points)
            far_columns = ["name", "address", "도로거리_km"]
            if is_season_far:
                far_columns += ["편도시간_분", "권장수행"]
                st.info("선택한 편도 시간 또는 거리 제한값을 넘는 대상입니다. "
                        "수행자·차량·제한 기준을 조정하거나 별도 순찰 대상으로 검토하세요.")
            st.dataframe(far_df[[column for column in far_columns if column in far_df.columns]],
                         use_container_width=True, hide_index=True)
            for p in far_points:
                st.markdown(
                    f"- [{p['name']} 길안내]({kakao_url(p['name'], p['lat'], p['lng'])})"
                    f" · 실도로거리 {p['도로거리_km']}km"
                    + (f" · 편도 약 {p.get('편도시간_분', '-')}분 · {p.get('권장수행', '')}"
                       if is_season_far else "")
                )

        st.divider()
        st.caption(
            "⚠️ 이 페이지의 API 키는 서버(Secrets)에만 저장되며 브라우저로 노출되지 않습니다. "
            "실제 도로거리·소요시간은 NCP Geocoding·Directions5 실시간 계산 결과입니다. "
            "📲 휴대폰에서는 브라우저 메뉴의 '홈 화면에 추가'를 누르면 앱처럼 사용할 수 있습니다."
        )

    st.info(
        "📌 본 앱이 생성한 순찰노선은 업무 지원을 위한 참고자료입니다. "
        "현장 여건, 도로 상황, 출동 공백 및 대상별 특성을 담당자가 충분히 검토한 후 사용하시기 바랍니다. "
        "최종 노선의 검토·결정과 운용 책임은 프로그램 사용자 및 해당 부서에 있습니다."
    )

    st.markdown(
        '<div style="margin-top:18px;padding:14px 8px 4px;text-align:center;'
        'border-top:1px solid #d9dee4;color:#5f6b64;font-size:0.86rem;">'
        '파세루 오리진&nbsp; | &nbsp;기획·개발 임미성&nbsp; | &nbsp;문의: '
        '<a href="mailto:emtmisung@gmail.com" style="color:#1f6fb2;'
        'text-decoration:none;">emtmisung@gmail.com</a></div>',
        unsafe_allow_html=True,
    )

